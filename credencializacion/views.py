# enrolamiento/views.py
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from .models import (
    Enrolamiento, SicreTblSig, EnrolamientoFamiliar, CargaMasiva,
    EnrolamientoCredencial, PlantillaCredencial, ConsecutivoFolio,
    UnidadAdministrativa, CODENAMES_CATALOGO_PERMISOS, AcuseCredencial,
)
from .serializers import (
    EnrolamientoSerializer, SigSerializer, EnrolamientoDataTableSerializer,
    ArchivoExcelSerializer, LoginSerializer, EnrolamientoFamiliarSerializer,
    CargaMasivaSerializer, EnrolamientoCredencialSerializer, PlantillaCredencialSerializer,
    UnidadAdministrativaSerializer, UsuarioSerializer, RolSerializer, PermisoSerializer,
    AcuseCredencialSerializer,
)
from .auth import EsSuperusuario, tiene_permiso
from . import media_utils
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.db.models import Q, Count, Min, Max
from django.db import models, transaction, connection
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
import pandas as pd
import base64
import os
import time
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from django.utils import timezone
import tempfile
import zipfile
from django.http import FileResponse


def _usuario_actual_id(request):
    """
    Id del usuario autenticado, para las columnas de auditoria
    (`id_usuario_registra` / `id_usuario_modifica`).

    Se toma de `request.user`, que es lo unico en lo que se puede confiar: un
    `id_usuario` mandado en el cuerpo lo elige el cliente y cualquiera podria
    atribuirle una impresion a otra persona -- justo lo que una auditoria no
    puede permitir.

    Regresa None si nadie inicio sesion, para que la columna quede vacia en
    vez de con un id inventado.
    """
    usuario = getattr(request, 'user', None)
    if usuario is not None and getattr(usuario, 'is_authenticated', False):
        return usuario.id
    return None


class AuditoriaUsuarioMixin:
    """
    Sella automaticamente quien creo/modifico la fila, en los modelos que
    tienen las columnas `id_usuario_registra` / `id_usuario_modifica`
    (EnrolamientoCredencial, PlantillaCredencial, Enrolamiento,
    EnrolamientoFamiliar, CargaMasiva).

    El check `hasattr` hace que este mixin sea seguro de poner en CUALQUIER
    ModelViewSet: en los modelos que no tienen esas columnas simplemente no
    hace nada.
    """
    def perform_create(self, serializer):
        campos = {f.name for f in self.get_queryset().model._meta.get_fields()}
        extra = {}
        if 'id_usuario_registra' in campos:
            extra['id_usuario_registra'] = _usuario_actual_id(self.request)
        serializer.save(**extra)

    def perform_update(self, serializer):
        campos = {f.name for f in self.get_queryset().model._meta.get_fields()}
        extra = {}
        if 'id_usuario_modifica' in campos:
            extra['id_usuario_modifica'] = _usuario_actual_id(self.request)
        serializer.save(**extra)


def _cargo_a_nivel(cargo: str) -> str:
    """
    Determina el nivel de credencial (nivel_credencial) a partir del campo CARGO/PUESTO.
    Niveles vÃ¡lidos:
        TITULAR, DIRECTOR_GENERAL, DIRECTOR_CENTRAL, DIRECTOR_DE_AREA,
        SUBDIRECTOR, JEFE_DE_DEPARTAMENTO, ENLACE, SEGURIDAD_INSTITUCIONAL
    """
    if not cargo:
        return 'ENLACE'
    c = str(cargo).strip().upper()

    if 'DIRECTOR GENERAL' in c or 'ADMINISTRADOR GENERAL' in c or 'TITULAR' in c:
        return 'TITULAR'
    if 'DIRECTOR CENTRAL' in c:
        return 'DIRECTOR_CENTRAL'
    if c.startswith('DIRECTOR DE AREA') or c.startswith('DIRECTOR DE ÃREA'):
        return 'DIRECTOR_DE_AREA'
    if c.startswith('DIRECTOR'):
        return 'DIRECTOR_DE_AREA'
    if 'SUBDIRECTOR' in c:
        return 'SUBDIRECTOR'
    if 'JEFE DE DEPARTAMENTO' in c or 'JEFE DEL DEPARTAMENTO' in c:
        return 'JEFE_DE_DEPARTAMENTO'
    if 'ENLACE' in c:
        return 'ENLACE'
    if 'OPERATIVO' in c or 'SEGURIDAD INSTITUCIONAL' in c or 'SEGURIDAD_INSTITUCIONAL' in c:
        return 'SEGURIDAD_INSTITUCIONAL'
    return 'ENLACE'


def extraer_imagenes_de_excel(archivo):
    """
    VersiÃ³n 5.0 (Final): Soporte DUAL para FOTOS (Col L) y FIRMAS (Col M).
    Wrapper que delega en extraer_imagenes_de_excel_real.
    """
    return extraer_imagenes_de_excel_real(archivo)


def _get_num_empleados_con_foto_firma():
    """
    Devuelve un set de valores de num_empleado que ya tienen FOTO y FIRMA
    en safirho_db.NW_EMPL_FOTO_ANAM (en todas sus variantes de formato).
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT EMPLID FROM safirho_db.NW_EMPL_FOTO_ANAM
                WHERE FOTO IS NOT NULL AND LENGTH(FOTO) > 1
                  AND FIRMA IS NOT NULL AND LENGTH(FIRMA) > 1
                """
            )
            rows = cursor.fetchall()
        result = set()
        for (emplid,) in rows:
            emplid_str = str(emplid).strip()
            result.add(emplid_str)                          # original  (00020232019)
            result.add(emplid_str.lstrip('0') or '0')      # sin ceros  (20232019)
        return result
    except Exception:
        return set()


def extraer_imagenes_de_excel_real(archivo):
    """\n    VersiÃ³n 5.0 (Final): Soporte DUAL para FOTOS (Col L) y FIRMAS (Col M).
    Mantiene el nombre original para compatibilidad.
    """
    from openpyxl import load_workbook
    from io import BytesIO
    import zipfile
    import posixpath
    import xml.etree.ElementTree as ET
    
    print(f"--- INICIANDO EXTRACCIÃ“N DUAL: {archivo} ---")
    fotos_por_fila = {}
    firmas_por_fila = {}
    
    # --- INTENTO 1: MÃ‰TODO ESTÃNDAR (OpenPyXL) ---
    try:
        if hasattr(archivo, 'seek'): archivo.seek(0)
        # data_only=False es vital para ver las imÃ¡genes
        wb = load_workbook(archivo, data_only=False)
        ws = wb.active
        
        imagenes = getattr(ws, '_images', []) or getattr(ws, 'images', [])
        print(f"   -> MÃ©todo EstÃ¡ndar detectÃ³: {len(imagenes)} imÃ¡genes")

        for image in imagenes:
            try:
                # Coordenadas (Excel base 1)
                # .anchor._from.row es Ã­ndice 0, por eso sumamos 1
                row = image.anchor._from.row + 1
                col = image.anchor._from.col + 1 
                
                # Extraer bytes
                img_bytes = None
                if hasattr(image, '_data'):
                    img_bytes = image._data()
                elif hasattr(image, 'ref'):
                    buf = BytesIO()
                    image.ref.save(buf)
                    img_bytes = buf.getvalue()
                
                if img_bytes:
                    # LÃ³gica de Columnas: 
                    # Columna 12 = L (Foto)
                    # Columna 13 = M (Firma)
                    if col == 12: 
                        fotos_por_fila[row] = img_bytes
                    elif col == 13:
                        firmas_por_fila[row] = img_bytes
                        
            except Exception as e:
                print(f"   -> Error leyendo imagen estÃ¡ndar: {e}")
        
        wb.close()
    except Exception as e:
        print(f"   -> FallÃ³ mÃ©todo estÃ¡ndar: {e}")

    # --- INTENTO 2: PLAN B ROBUSTO (ZIP + DRAWING RELS) ---
    # Si el mÃ©todo estÃ¡ndar fallÃ³ y no detectÃ³ NADA, resolvemos anclas reales
    # desde drawing.xml y sus relaciones (NO por orden alfabÃ©tico de archivos).
    if not fotos_por_fila and not firmas_por_fila:
        print("âš ï¸ ACTIVANDO EXTRACCIÃ“N ROBUSTA DUAL (PLAN B ZIP + RELS)")
        try:
            if hasattr(archivo, 'seek'): archivo.seek(0)
            if zipfile.is_zipfile(archivo):
                with zipfile.ZipFile(archivo, 'r') as z:
                    ns_main = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    ns_rel = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
                    ns_pkg_rel = {'pr': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    ns_draw = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    }

                    def _norm_zip_path(base_dir, target):
                        # target puede venir absoluto (/xl/...) o relativo (../drawings/...)
                        if target.startswith('/'):
                            target = target[1:]
                        else:
                            target = posixpath.normpath(posixpath.join(base_dir, target))
                        return target

                    # 1) Localizar worksheet activo real a partir de workbook.xml
                    wb_root = ET.fromstring(z.read('xl/workbook.xml'))
                    sheets = wb_root.find('m:sheets', ns_main)
                    first_sheet = sheets.find('m:sheet', ns_main) if sheets is not None else None
                    if first_sheet is None:
                        raise ValueError('No se encontrÃ³ ninguna hoja en workbook.xml')
                    sheet_rid = first_sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')

                    wb_rels_root = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
                    rel_node = None
                    for rel in wb_rels_root.findall('pr:Relationship', ns_pkg_rel):
                        if rel.attrib.get('Id') == sheet_rid:
                            rel_node = rel
                            break
                    if rel_node is None:
                        raise ValueError('No se pudo resolver la relaciÃ³n de la hoja activa')

                    sheet_path = _norm_zip_path('xl', rel_node.attrib.get('Target', ''))
                    sheet_dir = posixpath.dirname(sheet_path)

                    # 2) Leer drawing r:id desde la hoja activa
                    sheet_root = ET.fromstring(z.read(sheet_path))
                    drawing_tag = sheet_root.find('m:drawing', {**ns_main, **ns_rel})
                    if drawing_tag is None:
                        print('   -> La hoja activa no tiene drawing asociado')
                    else:
                        drawing_rid = drawing_tag.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')

                        # 3) Resolver drawing.xml desde sheet rels
                        sheet_rels_path = posixpath.join(sheet_dir, '_rels', posixpath.basename(sheet_path) + '.rels')
                        sheet_rels_root = ET.fromstring(z.read(sheet_rels_path))

                        drawing_rel = None
                        for rel in sheet_rels_root.findall('pr:Relationship', ns_pkg_rel):
                            if rel.attrib.get('Id') == drawing_rid:
                                drawing_rel = rel
                                break

                        if drawing_rel is None:
                            print('   -> No se encontrÃ³ relaciÃ³n al drawing en la hoja activa')
                        else:
                            drawing_path = _norm_zip_path(sheet_dir, drawing_rel.attrib.get('Target', ''))
                            drawing_dir = posixpath.dirname(drawing_path)
                            drawing_rels_path = posixpath.join(drawing_dir, '_rels', posixpath.basename(drawing_path) + '.rels')

                            # 4) Mapa rId -> target_path (mÃ¡s tolerante)
                            drawing_rels_root = ET.fromstring(z.read(drawing_rels_path))
                            media_by_rid = {}
                            for rel in drawing_rels_root.findall('pr:Relationship', ns_pkg_rel):
                                rel_id = rel.attrib.get('Id')
                                target = rel.attrib.get('Target', '')
                                if rel_id and target:
                                    media_by_rid[rel_id] = _norm_zip_path(drawing_dir, target)

                            print(f"   -> Relaciones en drawing: {len(media_by_rid)}")

                            # 5) Recorrer anchors con posiciÃ³n real y resolver bytes
                            drawing_root = ET.fromstring(z.read(drawing_path))
                            anchors = (
                                drawing_root.findall('.//xdr:twoCellAnchor', ns_draw)
                                + drawing_root.findall('.//xdr:oneCellAnchor', ns_draw)
                                + drawing_root.findall('.//xdr:absoluteAnchor', ns_draw)
                            )
                            print(f"   -> Anchors detectados en drawing: {len(anchors)}")

                            # Guardamos candidatos para asignar por fila/posiciÃ³n horizontal
                            candidates = []

                            for anchor in anchors:
                                from_node = anchor.find('xdr:from', ns_draw)
                                pic_node = anchor.find('.//xdr:pic', ns_draw)
                                if from_node is None or pic_node is None:
                                    continue

                                col_node = from_node.find('xdr:col', ns_draw)
                                row_node = from_node.find('xdr:row', ns_draw)
                                if col_node is None or row_node is None:
                                    continue

                                col = int(col_node.text) + 1
                                row = int(row_node.text) + 1

                                blip = pic_node.find('.//a:blip', ns_draw)
                                if blip is None:
                                    continue

                                rid_embed = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                rid_link = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}link')
                                rid = rid_embed or rid_link
                                if not rid:
                                    continue

                                media_path = media_by_rid.get(rid)
                                if not media_path:
                                    continue

                                # Solo nos interesan recursos internos de media
                                if 'xl/media/' not in media_path.replace('\\', '/'):
                                    continue

                                if media_path not in z.namelist():
                                    continue

                                img_bytes = z.read(media_path)
                                if not img_bytes:
                                    continue

                                candidates.append((row, col, media_path, img_bytes))

                            print(f"   -> Candidatos de imagen resueltos: {len(candidates)}")

                            # 6) Asignar foto/firma por fila usando orden horizontal (izq->der)
                            # Esto evita depender de si Excel ancla en K/L o L/M.
                            by_row = {}
                            for row, col, media_path, img_bytes in candidates:
                                by_row.setdefault(row, []).append((col, media_path, img_bytes))

                            for row, items in by_row.items():
                                items.sort(key=lambda x: x[0])

                                if len(items) == 1:
                                    col, media_path, img_bytes = items[0]
                                    # regla simple por columna aproximada
                                    if col <= 12:
                                        fotos_por_fila[row] = img_bytes
                                        print(f"   -> [ZIP-RELS] {media_path} es FOTO de Fila {row} (col={col})")
                                    else:
                                        firmas_por_fila[row] = img_bytes
                                        print(f"   -> [ZIP-RELS] {media_path} es FIRMA de Fila {row} (col={col})")
                                else:
                                    # Izquierda = Foto, Derecha = Firma
                                    col_f, media_f, bytes_f = items[0]
                                    col_s, media_s, bytes_s = items[1]
                                    fotos_por_fila[row] = bytes_f
                                    firmas_por_fila[row] = bytes_s
                                    print(f"   -> [ZIP-RELS] {media_f} es FOTO de Fila {row} (col={col_f})")
                                    print(f"   -> [ZIP-RELS] {media_s} es FIRMA de Fila {row} (col={col_s})")

        except Exception as e:
            print(f"âŒ FallÃ³ Plan B Dual (RELS): {e}")

    # Rebobinar siempre al final para que Pandas no falle despuÃ©s
    if hasattr(archivo, 'seek'): archivo.seek(0)
    
    return fotos_por_fila, firmas_por_fila


def procesar_foto_desde_excel(valor_celda, fila_numero=None, imagenes_excel=None):
    """
    Procesa el valor de la columna 'foto' del Excel y lo convierte a bytes.
    Prioridad:
    1. Imagen incrustada en Excel (si existe)
    2. Ruta de archivo
    3. Datos en base64
    4. None/vacÃ­o
    """
    # Prioridad 1: Imagen incrustada en el Excel
    if imagenes_excel and fila_numero and fila_numero in imagenes_excel:
        return imagenes_excel[fila_numero]
    
    if pd.isnull(valor_celda) or valor_celda == '':
        return None
    
    valor_str = str(valor_celda).strip()
    
    # Prioridad 2: Es una ruta de archivo (probar con y sin normalizaciÃ³n)
    rutas_a_probar = [
        valor_str,
        os.path.abspath(valor_str),
        valor_str.replace('\\', '/'),
        valor_str.replace('/', '\\')
    ]
    
    for ruta in rutas_a_probar:
        if os.path.exists(ruta) and os.path.isfile(ruta):
            try:
                with open(ruta, 'rb') as f:
                    return f.read()
            except Exception as e:
                print(f"Error leyendo archivo {ruta}: {e}")
                continue
    
    # Prioridad 3: Es base64
    try:
        if 'base64,' in valor_str:
            valor_str = valor_str.split('base64,')[1]
        return base64.b64decode(valor_str)
    except Exception:
        return None


class EnrolamientoViewSet(AuditoriaUsuarioMixin, viewsets.ModelViewSet):
    queryset = Enrolamiento.objects.all().order_by('-id_enrolamiento')
    serializer_class = EnrolamientoSerializer
    
    filter_backends = [filters.SearchFilter]
    search_fields = ['rfc', 'num_empleado', 'nombre', 'paterno']   

    @action(detail=False, methods=['get'], url_path='listos-imprimir')
    def listos_para_imprimir(self, request):
        """
        Muestra solo los registros COMPLETOS:
        - Tienen Foto
        - Tienen Firma
        - Tienen NÃºmero de Empleado
        """
        queryset = self.get_queryset().exclude(
            Q(foto__isnull=True) | Q(foto='') |
            Q(firma__isnull=True) | Q(firma='') |
            Q(num_empleado__isnull=True) | Q(num_empleado='')
        )
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='datatableImprimir')
    def datatableImprimir(self, request):
        """
        Endpoint para alimentar la data table del front.
        Columnas: Num Empleado, RFC, CURP, nombres, adscripcion, puesto.
        Filtros: foto y firma no nulos (incluyendo flag 1 o registro en NW_EMPL_FOTO_ANAM),
        y que no estÃ©n marcados como impresos (impreso != 1).
        """
        completos = _get_num_empleados_con_foto_firma()

        queryset = self.get_queryset().exclude(
            Q(foto__isnull=True) | Q(foto=b'') |
            Q(firma__isnull=True) | Q(firma=b'')
        ).filter(
            Q(impreso__isnull=True) | Q(impreso=0) | ~Q(impreso=1)
        )

        # TambiÃ©n incluir empleados con foto+firma en NW_EMPL_FOTO_ANAM
        if completos:
            qs_externos = self.get_queryset().filter(
                num_empleado__in=completos
            ).filter(
                Q(impreso__isnull=True) | Q(impreso=0) | ~Q(impreso=1)
            )
            from itertools import chain
            queryset = list(chain(queryset, qs_externos.exclude(pk__in=queryset.values_list('pk', flat=True))))
        else:
            queryset = list(queryset)

        queryset_familiares = EnrolamientoFamiliar.objects.exclude(
            Q(foto__isnull=True) | Q(foto=b'') |
            Q(firma__isnull=True) | Q(firma=b'')
        ).filter(
            Q(impreso__isnull=True) | Q(impreso=0) | ~Q(impreso=1)
        )

        serializer_enrolamiento = EnrolamientoDataTableSerializer(queryset, many=True)
        serializer_familiar = EnrolamientoFamiliarSerializer(queryset_familiares, many=True)

        data_enrolamiento = []
        for item in serializer_enrolamiento.data:
            registro = dict(item)
            registro['source_table'] = 'enrolamiento'
            if int(registro.get('nuevo_laredo') or 0) == 1:
                registro['tipo_credencial'] = 'provisional'
            else:
                registro['tipo_credencial'] = 'anam'
            data_enrolamiento.append(registro)

        data_familiares = []
        for item in serializer_familiar.data:
            registro = dict(item)
            registro['source_table'] = 'familiar'
            registro['tipo_credencial'] = 'familiar'
            registro['folio'] = registro.get('folio_familiares')
            data_familiares.append(registro)

        data = sorted(
            data_enrolamiento + data_familiares,
            key=lambda x: x.get('fecha_enrolamiento') or x.get('fecha_registro') or '',
            reverse=True
        )
        
        page = self.paginate_queryset(data)
        if page is not None:
            return self.get_paginated_response(page)
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def pendientes(self, request):
        """
        Muestra solo los registros que faltan de foto O firma,
        y que adicionalmente NO tienen ambas en safirho_db.NW_EMPL_FOTO_ANAM.
        """
        completos = _get_num_empleados_con_foto_firma()

        queryset = self.get_queryset().filter(
            Q(foto__isnull=True) | Q(foto=b'') |
            Q(firma__isnull=True) | Q(firma=b'')
        )

        if completos:
            queryset = queryset.exclude(num_empleado__in=completos)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get', 'post'])
    def sincronizar(self, request):
        """
        Copia los datos de SicreTblSig a Enrolamiento.
        """
        registros_sig = SicreTblSig.objects.all()
        creados = 0
        actualizados = 0

        for sig in registros_sig:
            nombre = (sig.nombres or '').strip()
            paterno = (sig.primer_apellido or '').strip()
            materno = (sig.segundo_apellido or '').strip()
            apellidos = f"{paterno} {materno}".strip()
            rfc_ref = (sig.empleado_anam or sig.no_empleado or sig.curp or '').strip()

            obj, created = Enrolamiento.objects.update_or_create(
                curp=sig.curp,
                defaults={
                    'rfc': rfc_ref,
                    'num_empleado': sig.no_empleado,
                    'curp': sig.curp,
                    'nombre': nombre,
                    'paterno': paterno,
                    'materno': materno,
                    'apellidos': apellidos,
                    'puesto': sig.cargo,
                    'adscripcion': sig.area,
                    'activo': 1
                }
            )
          
            if not created:
                obj.save() 

            if created:
                creados += 1
            else:
                actualizados += 1

            

        return Response({
            'status': 'SincronizaciÃ³n completada',
            'creados': creados,
            'actualizados': actualizados
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='obtener-folio-maximo')
    def obtener_folio_maximo(self, request):
        """
        Retorna el folio mÃ¡ximo actual de la tabla enrolamiento.
        El front puede usar este valor para asignar el siguiente consecutivo.
        Respeta el formato con ceros a la izquierda (ej: 000001 -> 000002).
        Permite separar series por nuevo_laredo con query param ?nuevo_laredo=0|1
        """
        try:
            nuevo_laredo_param = request.query_params.get('nuevo_laredo', None)
            folios_queryset = Enrolamiento.objects.exclude(
                Q(folio__isnull=True) | Q(folio='')
            )

            if nuevo_laredo_param in ['0', '1']:
                nuevo_laredo_flag = int(nuevo_laredo_param)
                if nuevo_laredo_flag == 1:
                    folios_queryset = folios_queryset.filter(nuevo_laredo=1)
                else:
                    folios_queryset = folios_queryset.filter(
                        Q(nuevo_laredo=0) | Q(nuevo_laredo__isnull=True)
                    )

            # Filtrar folios no nulos y obtener todos
            folios = folios_queryset.values_list('folio', flat=True)
            
            if not folios:
                # Si no hay folios, empezar desde 000001 (6 dÃ­gitos por defecto)
                return Response({
                    'status': 'success',
                    'folio_maximo': '000000',
                    'siguiente_folio': '000001',
                    'serie_nuevo_laredo': nuevo_laredo_param
                }, status=status.HTTP_200_OK)
            
            # Encontrar el folio con mayor valor numÃ©rico
            folio_max_valor = 0
            folio_max_str = ''
            longitud_formato = 6  # Default
            
            for folio in folios:
                folio_str = str(folio).strip()
                try:
                    valor_numerico = int(folio_str)
                    if valor_numerico > folio_max_valor:
                        folio_max_valor = valor_numerico
                        folio_max_str = folio_str
                        longitud_formato = len(folio_str)
                except (ValueError, TypeError):
                    continue
            
            if folio_max_valor == 0:
                # No se encontraron folios vÃ¡lidos
                return Response({
                    'status': 'success',
                    'folio_maximo': '000000',
                    'siguiente_folio': '000001',
                    'serie_nuevo_laredo': nuevo_laredo_param
                }, status=status.HTTP_200_OK)
            
            # Calcular siguiente folio manteniendo el formato
            siguiente_valor = folio_max_valor + 1
            siguiente_folio = str(siguiente_valor).zfill(longitud_formato)
            
            return Response({
                'status': 'success',
                'folio_maximo': folio_max_str,
                'siguiente_folio': siguiente_folio,
                'serie_nuevo_laredo': nuevo_laredo_param
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al obtener folio mÃ¡ximo: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='guardar-foto-firma')
    def guardar_foto_firma(self, request, pk=None):
        """
        Guarda foto y firma en safirho_db.NW_EMPL_FOTO_ANAM (INSERT o UPDATE).
        Marca enrolamiento.foto y enrolamiento.firma con flag b'\\x01' para excluirlo
        del endpoint pendientes sin guardar el blob en sicre_db.
        """
        enrolamiento = self.get_object()
        foto_b64 = request.data.get('foto')
        firma_b64 = request.data.get('firma')

        if not foto_b64 or not firma_b64:
            return Response({'status': 'error', 'mensaje': 'Faltan foto o firma'}, status=400)

        def clean_b64(data):
            if isinstance(data, str) and 'base64,' in data:
                data = data.split('base64,')[1]
            return base64.b64decode(data)

        try:
            foto_bytes = clean_b64(foto_b64)
            firma_bytes = clean_b64(firma_b64)
        except Exception as e:
            return Response({'status': 'error', 'mensaje': f'Error al decodificar imagen: {e}'}, status=400)

        emplid = str(enrolamiento.num_empleado or '').strip()
        if not emplid:
            return Response({'status': 'error', 'mensaje': 'El empleado no tiene nÃºmero de empleado'}, status=400)

        emplid_padded = emplid.zfill(11)
        emplid_stripped = emplid.lstrip('0') or '0'
        variantes = list({emplid, emplid_padded, emplid_stripped})
        placeholders = ','.join(['%s'] * len(variantes))

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    f"SELECT EMPLID FROM safirho_db.NW_EMPL_FOTO_ANAM WHERE EMPLID IN ({placeholders}) LIMIT 1",
                    variantes
                )
                existing = cursor.fetchone()
                if existing:
                    cursor.execute(
                        "UPDATE safirho_db.NW_EMPL_FOTO_ANAM "
                        "SET FOTO=%s, FIRMA=%s, TipoMime='image/jpeg', updated_at=NOW() "
                        "WHERE EMPLID=%s",
                        [foto_bytes, firma_bytes, existing[0]]
                    )
                else:
                    cursor.execute(
                        "INSERT INTO safirho_db.NW_EMPL_FOTO_ANAM "
                        "(EMPLID, FOTO, FIRMA, TipoMime, updated_at) VALUES (%s, %s, %s, 'image/jpeg', NOW())",
                        [emplid_padded, foto_bytes, firma_bytes]
                    )
        except Exception as e:
            return Response({'status': 'error', 'mensaje': f'Error al guardar en BD externa: {e}'}, status=500)

        enrolamiento.foto = b'\x01'
        enrolamiento.firma = b'\x01'
        enrolamiento.save(update_fields=['foto', 'firma'])

        return Response({'status': 'success', 'mensaje': 'Foto y firma guardadas correctamente'})

    @action(detail=False, methods=['get', 'post'], url_path='pendientes-de-imprimir')
    def pendientes_de_imprimir(self, request):
        """
        Endpoint de bÃºsqueda avanzada para ag-grid.
        Permite filtrar por mÃºltiples campos y retorna todos los detalles.
        Soporta filtros por: num_empleado, rfc, curp, nombre, paterno, materno,
        puesto, adscripcion, folio, impreso, fecha_expedicion_desde, fecha_expedicion_hasta.
        """
        queryset = self.get_queryset().filter(fecha_expedicion__isnull=False).filter(Q(impreso=0) | Q(impreso__isnull=True))
        queryset_familiares = EnrolamientoFamiliar.objects.filter(fecha_expedicion__isnull=False).filter(Q(impreso=0) | Q(impreso__isnull=True)).order_by('-id_enrolamiento')
        
        # Obtener parÃ¡metros de filtro (soporta GET y POST)
        params = request.query_params if request.method == 'GET' else request.data
        
        # Filtros de texto
        if params.get('num_empleado'):
            queryset = queryset.filter(num_empleado__icontains=params['num_empleado'])
            queryset_familiares = queryset_familiares.filter(num_empleado__icontains=params['num_empleado'])
        
        if params.get('rfc'):
            queryset = queryset.filter(rfc__icontains=params['rfc'])
            queryset_familiares = queryset_familiares.filter(rfc__icontains=params['rfc'])
        
        if params.get('curp'):
            queryset = queryset.filter(curp__icontains=params['curp'])
            queryset_familiares = queryset_familiares.filter(curp__icontains=params['curp'])
        
        if params.get('nombre'):
            queryset = queryset.filter(nombre__icontains=params['nombre'])
            queryset_familiares = queryset_familiares.filter(nombre__icontains=params['nombre'])
        
        if params.get('paterno'):
            queryset = queryset.filter(paterno__icontains=params['paterno'])
            queryset_familiares = queryset_familiares.filter(paterno__icontains=params['paterno'])
        
        if params.get('materno'):
            queryset = queryset.filter(materno__icontains=params['materno'])
            queryset_familiares = queryset_familiares.filter(materno__icontains=params['materno'])
        
        if params.get('puesto'):
            queryset = queryset.filter(puesto__icontains=params['puesto'])
            queryset_familiares = queryset_familiares.filter(puesto__icontains=params['puesto'])
        
        if params.get('adscripcion'):
            queryset = queryset.filter(adscripcion__icontains=params['adscripcion'])
            queryset_familiares = queryset_familiares.filter(adscripcion__icontains=params['adscripcion'])
        
        if params.get('folio'):
            queryset = queryset.filter(folio__icontains=params['folio'])
            queryset_familiares = queryset_familiares.filter(folio_familiares__icontains=params['folio'])
        
        # Filtro por estado de impresiÃ³n
        if params.get('impreso') is not None:
            impreso_val = params['impreso']
            if impreso_val == '1' or impreso_val == 1:
                queryset = queryset.filter(impreso=1)
                queryset_familiares = queryset_familiares.filter(impreso=1)
            elif impreso_val == '0' or impreso_val == 0:
                queryset = queryset.filter(Q(impreso=0) | Q(impreso__isnull=True))
                queryset_familiares = queryset_familiares.filter(Q(impreso=0) | Q(impreso__isnull=True))
        
        # Filtros por rango de fechas de expediciÃ³n
        if params.get('fecha_expedicion_desde'):
            queryset = queryset.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
        
        if params.get('fecha_expedicion_hasta'):
            queryset = queryset.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
        
        # Filtros por rango de fechas de vigencia
        if params.get('inicio_vig_desde'):
            queryset = queryset.filter(inicio_vig__gte=params['inicio_vig_desde'])
            queryset_familiares = queryset_familiares.filter(inicio_vig__gte=params['inicio_vig_desde'])
        
        if params.get('fin_vig_hasta'):
            queryset = queryset.filter(fin_vig__lte=params['fin_vig_hasta'])
            queryset_familiares = queryset_familiares.filter(fin_vig__lte=params['fin_vig_hasta'])
        
        # Filtros por fecha de registro
        if params.get('fecha_registro_desde'):
            queryset = queryset.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
        
        if params.get('fecha_registro_hasta'):
            queryset = queryset.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
        
        # Filtro por fecha de registro especÃ­fica
        if params.get('fecha_registro'):
            queryset = queryset.filter(fecha_registro__date=params['fecha_registro'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date=params['fecha_registro'])
        
        # Filtro por fecha de expediciÃ³n especÃ­fica
        if params.get('fecha_expedicion'):
            queryset = queryset.filter(fecha_expedicion=params['fecha_expedicion'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion=params['fecha_expedicion'])
        
        # Filtro por estado de completitud (con foto y firma)
        if params.get('solo_completos') == 'true' or params.get('solo_completos') == True:
            queryset = queryset.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
            queryset_familiares = queryset_familiares.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
        
        # Filtro por registros activos
        if params.get('solo_activos') == 'true' or params.get('solo_activos') == True:
            queryset = queryset.filter(activo=1)
            queryset_familiares = queryset_familiares.filter(activo=1)

        serializer_enrolamiento = self.get_serializer(queryset, many=True)
        serializer_familiar = EnrolamientoFamiliarSerializer(queryset_familiares, many=True)

        data_enrolamiento = []
        for item in serializer_enrolamiento.data:
            registro = dict(item)
            registro['source_table'] = 'enrolamiento'
            if int(registro.get('nuevo_laredo') or 0) == 1:
                registro['tipo_credencial'] = 'provisional'
            elif int(registro.get('provisional') or 0) == 1:
                registro['tipo_credencial'] = 'anam'
            else:
                registro['tipo_credencial'] = 'enrolamiento'
            data_enrolamiento.append(registro)

        data_familiares = []
        for item in serializer_familiar.data:
            registro = dict(item)
            registro['source_table'] = 'familiar'
            registro['tipo_credencial'] = 'familiar'
            registro['folio'] = registro.get('folio_familiares')
            data_familiares.append(registro)

        data = sorted(
            data_enrolamiento + data_familiares,
            key=lambda x: x.get('fecha_registro') or x.get('fecha_enrolamiento') or '',
            reverse=True
        )

        page = self.paginate_queryset(data)
        if page is not None:
            return self.get_paginated_response(page)

        return Response(data)

    @action(detail=False, methods=['get', 'post'], url_path='pendientes-de-imprimir')
    def pendientes_de_imprimir(self, request):
        """
        Endpoint de bÃºsqueda avanzada para ag-grid.
        Permite filtrar por mÃºltiples campos y retorna todos los detalles.
        Soporta filtros por: num_empleado, rfc, curp, nombre, paterno, materno,
        puesto, adscripcion, folio, impreso, fecha_expedicion_desde, fecha_expedicion_hasta.
        """
        queryset = self.get_queryset().filter(fecha_expedicion__isnull=False).filter(Q(impreso=0) | Q(impreso__isnull=True))
        queryset_familiares = EnrolamientoFamiliar.objects.filter(fecha_expedicion__isnull=False).filter(Q(impreso=0) | Q(impreso__isnull=True)).order_by('-id_enrolamiento')
        
        # Obtener parÃ¡metros de filtro (soporta GET y POST)
        params = request.query_params if request.method == 'GET' else request.data
        
        # Filtros de texto
        if params.get('num_empleado'):
            queryset = queryset.filter(num_empleado__icontains=params['num_empleado'])
            queryset_familiares = queryset_familiares.filter(num_empleado__icontains=params['num_empleado'])
        
        if params.get('rfc'):
            queryset = queryset.filter(rfc__icontains=params['rfc'])
            queryset_familiares = queryset_familiares.filter(rfc__icontains=params['rfc'])
        
        if params.get('curp'):
            queryset = queryset.filter(curp__icontains=params['curp'])
            queryset_familiares = queryset_familiares.filter(curp__icontains=params['curp'])
        
        if params.get('nombre'):
            queryset = queryset.filter(nombre__icontains=params['nombre'])
            queryset_familiares = queryset_familiares.filter(nombre__icontains=params['nombre'])
        
        if params.get('paterno'):
            queryset = queryset.filter(paterno__icontains=params['paterno'])
            queryset_familiares = queryset_familiares.filter(paterno__icontains=params['paterno'])
        
        if params.get('materno'):
            queryset = queryset.filter(materno__icontains=params['materno'])
            queryset_familiares = queryset_familiares.filter(materno__icontains=params['materno'])
        
        if params.get('puesto'):
            queryset = queryset.filter(puesto__icontains=params['puesto'])
            queryset_familiares = queryset_familiares.filter(puesto__icontains=params['puesto'])
        
        if params.get('adscripcion'):
            queryset = queryset.filter(adscripcion__icontains=params['adscripcion'])
            queryset_familiares = queryset_familiares.filter(adscripcion__icontains=params['adscripcion'])
        
        if params.get('folio'):
            queryset = queryset.filter(folio__icontains=params['folio'])
            queryset_familiares = queryset_familiares.filter(folio_familiares__icontains=params['folio'])
        
        # Filtro por estado de impresiÃ³n
        if params.get('impreso') is not None:
            impreso_val = params['impreso']
            if impreso_val == '1' or impreso_val == 1:
                queryset = queryset.filter(impreso=1)
                queryset_familiares = queryset_familiares.filter(impreso=1)
            elif impreso_val == '0' or impreso_val == 0:
                queryset = queryset.filter(Q(impreso=0) | Q(impreso__isnull=True))
                queryset_familiares = queryset_familiares.filter(Q(impreso=0) | Q(impreso__isnull=True))
        
        # Filtros por rango de fechas de expediciÃ³n
        if params.get('fecha_expedicion_desde'):
            queryset = queryset.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
        
        if params.get('fecha_expedicion_hasta'):
            queryset = queryset.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
        
        # Filtros por rango de fechas de vigencia
        if params.get('inicio_vig_desde'):
            queryset = queryset.filter(inicio_vig__gte=params['inicio_vig_desde'])
            queryset_familiares = queryset_familiares.filter(inicio_vig__gte=params['inicio_vig_desde'])
        
        if params.get('fin_vig_hasta'):
            queryset = queryset.filter(fin_vig__lte=params['fin_vig_hasta'])
            queryset_familiares = queryset_familiares.filter(fin_vig__lte=params['fin_vig_hasta'])
        
        # Filtros por fecha de registro
        if params.get('fecha_registro_desde'):
            queryset = queryset.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
        
        if params.get('fecha_registro_hasta'):
            queryset = queryset.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
        
        # Filtro por fecha de registro especÃ­fica
        if params.get('fecha_registro'):
            queryset = queryset.filter(fecha_registro__date=params['fecha_registro'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date=params['fecha_registro'])
        
        # Filtro por fecha de expediciÃ³n especÃ­fica
        if params.get('fecha_expedicion'):
            queryset = queryset.filter(fecha_expedicion=params['fecha_expedicion'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion=params['fecha_expedicion'])
        
        # Filtro por estado de completitud (con foto y firma)
        if params.get('solo_completos') == 'true' or params.get('solo_completos') == True:
            queryset = queryset.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
            queryset_familiares = queryset_familiares.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
        
        # Filtro por registros activos
        if params.get('solo_activos') == 'true' or params.get('solo_activos') == True:
            queryset = queryset.filter(activo=1)
            queryset_familiares = queryset_familiares.filter(activo=1)

        serializer_enrolamiento = self.get_serializer(queryset, many=True)
        serializer_familiar = EnrolamientoFamiliarSerializer(queryset_familiares, many=True)

        data_enrolamiento = []
        for item in serializer_enrolamiento.data:
            registro = dict(item)
            registro['source_table'] = 'enrolamiento'
            if int(registro.get('nuevo_laredo') or 0) == 1:
                registro['tipo_credencial'] = 'provisional'
            elif int(registro.get('provisional') or 0) == 1:
                registro['tipo_credencial'] = 'anam'
            else:
                registro['tipo_credencial'] = 'enrolamiento'
            data_enrolamiento.append(registro)

        data_familiares = []
        for item in serializer_familiar.data:
            registro = dict(item)
            registro['source_table'] = 'familiar'
            registro['tipo_credencial'] = 'familiar'
            registro['folio'] = registro.get('folio_familiares')
            data_familiares.append(registro)

        data = sorted(
            data_enrolamiento + data_familiares,
            key=lambda x: x.get('fecha_registro') or x.get('fecha_enrolamiento') or '',
            reverse=True
        )

        page = self.paginate_queryset(data)
        if page is not None:
            return self.get_paginated_response(page)

        return Response(data)

    @action(detail=False, methods=['get', 'post'], url_path='busqueda-avanzada')
    def busqueda_avanzada(self, request):
        """
        Endpoint de bÃºsqueda avanzada para ag-grid.
        Permite filtrar por mÃºltiples campos y retorna todos los detalles.
        Soporta filtros por: num_empleado, rfc, curp, nombre, paterno, materno,
        puesto, adscripcion, folio, impreso, fecha_expedicion_desde, fecha_expedicion_hasta.
        """
        queryset = self.get_queryset()
        queryset_familiares = EnrolamientoFamiliar.objects.all().order_by('-id_enrolamiento')
        
        # Obtener parÃ¡metros de filtro (soporta GET y POST)
        params = request.query_params if request.method == 'GET' else request.data
        
        # Filtros de texto
        if params.get('num_empleado'):
            queryset = queryset.filter(num_empleado__icontains=params['num_empleado'])
            queryset_familiares = queryset_familiares.filter(num_empleado__icontains=params['num_empleado'])
        
        if params.get('rfc'):
            queryset = queryset.filter(rfc__icontains=params['rfc'])
            queryset_familiares = queryset_familiares.filter(rfc__icontains=params['rfc'])
        
        if params.get('curp'):
            queryset = queryset.filter(curp__icontains=params['curp'])
            queryset_familiares = queryset_familiares.filter(curp__icontains=params['curp'])
        
        if params.get('nombre'):
            queryset = queryset.filter(nombre__icontains=params['nombre'])
            queryset_familiares = queryset_familiares.filter(nombre__icontains=params['nombre'])
        
        if params.get('paterno'):
            queryset = queryset.filter(paterno__icontains=params['paterno'])
            queryset_familiares = queryset_familiares.filter(paterno__icontains=params['paterno'])
        
        if params.get('materno'):
            queryset = queryset.filter(materno__icontains=params['materno'])
            queryset_familiares = queryset_familiares.filter(materno__icontains=params['materno'])
        
        if params.get('puesto'):
            queryset = queryset.filter(puesto__icontains=params['puesto'])
            queryset_familiares = queryset_familiares.filter(puesto__icontains=params['puesto'])
        
        if params.get('adscripcion'):
            queryset = queryset.filter(adscripcion__icontains=params['adscripcion'])
            queryset_familiares = queryset_familiares.filter(adscripcion__icontains=params['adscripcion'])
        
        if params.get('folio'):
            queryset = queryset.filter(folio__icontains=params['folio'])
            queryset_familiares = queryset_familiares.filter(folio_familiares__icontains=params['folio'])
        
        # Filtro por estado de impresiÃ³n
        if params.get('impreso') is not None:
            impreso_val = params['impreso']
            if impreso_val == '1' or impreso_val == 1:
                queryset = queryset.filter(impreso=1)
                queryset_familiares = queryset_familiares.filter(impreso=1)
            elif impreso_val == '0' or impreso_val == 0:
                queryset = queryset.filter(Q(impreso=0) | Q(impreso__isnull=True))
                queryset_familiares = queryset_familiares.filter(Q(impreso=0) | Q(impreso__isnull=True))
        
        # Filtros por rango de fechas de expediciÃ³n
        if params.get('fecha_expedicion_desde'):
            queryset = queryset.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__gte=params['fecha_expedicion_desde'])
        
        if params.get('fecha_expedicion_hasta'):
            queryset = queryset.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion__lte=params['fecha_expedicion_hasta'])
        
        # Filtros por rango de fechas de vigencia
        if params.get('inicio_vig_desde'):
            queryset = queryset.filter(inicio_vig__gte=params['inicio_vig_desde'])
            queryset_familiares = queryset_familiares.filter(inicio_vig__gte=params['inicio_vig_desde'])
        
        if params.get('fin_vig_hasta'):
            queryset = queryset.filter(fin_vig__lte=params['fin_vig_hasta'])
            queryset_familiares = queryset_familiares.filter(fin_vig__lte=params['fin_vig_hasta'])
        
        # Filtros por fecha de registro
        if params.get('fecha_registro_desde'):
            queryset = queryset.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__gte=params['fecha_registro_desde'])
        
        if params.get('fecha_registro_hasta'):
            queryset = queryset.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date__lte=params['fecha_registro_hasta'])
        
        # Filtro por fecha de registro especÃ­fica
        if params.get('fecha_registro'):
            queryset = queryset.filter(fecha_registro__date=params['fecha_registro'])
            queryset_familiares = queryset_familiares.filter(fecha_registro__date=params['fecha_registro'])
        
        # Filtro por fecha de expediciÃ³n especÃ­fica
        if params.get('fecha_expedicion'):
            queryset = queryset.filter(fecha_expedicion=params['fecha_expedicion'])
            queryset_familiares = queryset_familiares.filter(fecha_expedicion=params['fecha_expedicion'])
        
        # Filtro por estado de completitud (con foto y firma)
        if params.get('solo_completos') == 'true' or params.get('solo_completos') == True:
            queryset = queryset.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
            queryset_familiares = queryset_familiares.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            )
        
        # Filtro por registros activos
        if params.get('solo_activos') == 'true' or params.get('solo_activos') == True:
            queryset = queryset.filter(activo=1)
            queryset_familiares = queryset_familiares.filter(activo=1)

        serializer_enrolamiento = self.get_serializer(queryset, many=True)
        serializer_familiar = EnrolamientoFamiliarSerializer(queryset_familiares, many=True)

        data_enrolamiento = []
        for item in serializer_enrolamiento.data:
            registro = dict(item)
            registro['source_table'] = 'enrolamiento'
            if int(registro.get('nuevo_laredo') or 0) == 1:
                registro['tipo_credencial'] = 'provisional'
            elif int(registro.get('provisional') or 0) == 1:
                registro['tipo_credencial'] = 'anam'
            else:
                registro['tipo_credencial'] = 'enrolamiento'
            data_enrolamiento.append(registro)

        data_familiares = []
        for item in serializer_familiar.data:
            registro = dict(item)
            registro['source_table'] = 'familiar'
            registro['tipo_credencial'] = 'familiar'
            registro['folio'] = registro.get('folio_familiares')
            data_familiares.append(registro)

        data = sorted(
            data_enrolamiento + data_familiares,
            key=lambda x: x.get('fecha_registro') or x.get('fecha_enrolamiento') or '',
            reverse=True
        )

        page = self.paginate_queryset(data)
        if page is not None:
            return self.get_paginated_response(page)

        return Response(data)
    
    @action(detail=False, methods=['get'], url_path='estadisticas')
    def estadisticas(self, request):
        """
        Retorna estadÃ­sticas generales del sistema de credencializaciÃ³n.
        Incluye: total de credenciales, impresas, pendientes, impresas hoy,
        por adscripciÃ³n, y tendencias.
        Soporta filtrado por rango de fechas (fecha_desde, fecha_hasta).
        """
        from django.db.models import Count, Q
        from datetime import datetime, timedelta
        
        try:
            # Fecha de hoy
            hoy = datetime.now().date()
            
            # Obtener parÃ¡metros de filtrado por fechas
            fecha_desde = request.query_params.get('fecha_desde')
            fecha_hasta = request.query_params.get('fecha_hasta')
            
            # Crear queryset base con filtros opcionales
            queryset_base = Enrolamiento.objects.all()
            if fecha_desde:
                queryset_base = queryset_base.filter(fecha_registro__date__gte=fecha_desde)
            if fecha_hasta:
                queryset_base = queryset_base.filter(fecha_registro__date__lte=fecha_hasta)
            
            # Total de registros
            total_credenciales = queryset_base.count()
            
            # Credenciales completas (con foto y firma)
            credenciales_completas = queryset_base.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            ).count()
            
            # Credenciales impresas
            credenciales_impresas = queryset_base.filter(impreso=1).count()
            
            # Credenciales impresas hoy
            credenciales_impresas_hoy = queryset_base.filter(
                impreso=1,
                fecha_expedicion=hoy
            ).count()
            
            # Credenciales pendientes de impresiÃ³n (completas pero no impresas)
            credenciales_pendientes = queryset_base.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            ).filter(
                Q(impreso__isnull=True) | Q(impreso=0) | ~Q(impreso=1)
            ).count()
            
            # Lista detallada de pendientes (para mostrar al dar click en el banner)
            lista_pendientes = queryset_base.exclude(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            ).filter(
                Q(impreso__isnull=True) | Q(impreso=0) | ~Q(impreso=1)
            ).values('id_enrolamiento', 'num_empleado', 'rfc', 'nombre', 'paterno', 'materno', 'folio', 'adscripcion')
            
            # Credenciales incompletas (sin foto o sin firma)
            credenciales_incompletas = queryset_base.filter(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            ).count()
            
            # Lista detallada de incompletas
            lista_incompletas = queryset_base.filter(
                Q(foto__isnull=True) | Q(foto=b'') |
                Q(firma__isnull=True) | Q(firma=b'')
            ).values('id_enrolamiento', 'num_empleado', 'rfc', 'nombre', 'paterno', 'materno', 'adscripcion')
            
            # Credenciales por adscripciÃ³n (top 10)
            por_adscripcion = queryset_base.values('adscripcion').annotate(
                total=Count('id_enrolamiento')
            ).order_by('-total')[:10]
            
            # EstadÃ­sticas de los Ãºltimos 7 dÃ­as
            hace_7_dias = hoy - timedelta(days=7)
            credenciales_ultima_semana = queryset_base.filter(
                fecha_registro__gte=hace_7_dias
            ).count()
            
            impresas_ultima_semana = queryset_base.filter(
                impreso=1,
                fecha_expedicion__gte=hace_7_dias
            ).count()
            
            # EstadÃ­sticas del mes actual
            inicio_mes = hoy.replace(day=1)
            credenciales_mes_actual = queryset_base.filter(
                fecha_registro__gte=inicio_mes
            ).count()
            
            impresas_mes_actual = queryset_base.filter(
                impreso=1,
                fecha_expedicion__gte=inicio_mes
            ).count()
            
            # Porcentajes
            porcentaje_impresas = round((credenciales_impresas / total_credenciales * 100), 2) if total_credenciales > 0 else 0
            porcentaje_completas = round((credenciales_completas / total_credenciales * 100), 2) if total_credenciales > 0 else 0
            porcentaje_pendientes = round((credenciales_pendientes / credenciales_completas * 100), 2) if credenciales_completas > 0 else 0
            
            return Response({
                'status': 'success',
                'totales': {
                    'total_credenciales': total_credenciales,
                    'credenciales_completas': credenciales_completas,
                    'credenciales_impresas': credenciales_impresas,
                    'credenciales_pendientes': credenciales_pendientes,
                    'credenciales_incompletas': credenciales_incompletas
                },
                'hoy': {
                    'credenciales_impresas_hoy': credenciales_impresas_hoy
                },
                'ultima_semana': {
                    'credenciales_registradas': credenciales_ultima_semana,
                    'credenciales_impresas': impresas_ultima_semana
                },
                'mes_actual': {
                    'credenciales_registradas': credenciales_mes_actual,
                    'credenciales_impresas': impresas_mes_actual
                },
                'porcentajes': {
                    'porcentaje_impresas': porcentaje_impresas,
                    'porcentaje_completas': porcentaje_completas,
                    'porcentaje_pendientes': porcentaje_pendientes
                },
                'por_adscripcion': list(por_adscripcion),
                'detalle_pendientes': list(lista_pendientes),
                'detalle_incompletas': list(lista_incompletas)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al obtener estadísticas: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='marcar-impreso')
    def marcar_impreso(self, request, pk=None):
        try:
            enrolamiento = self.get_object()
            
            enrolamiento.impreso = 1
            fecha_expedicion_payload = request.data.get('fecha_expedicion')
            if not enrolamiento.fecha_expedicion:
                from django.utils import timezone
                enrolamiento.fecha_expedicion = fecha_expedicion_payload or timezone.now().date()
            enrolamiento.save()

            return Response({
                'status': 'success',
                'mensaje': f'Credencial empleada {enrolamiento.folio or enrolamiento.id_enrolamiento} marcada como impresa',
                'data': {
                    'id_enrolamiento': enrolamiento.id_enrolamiento,
                    'folio': enrolamiento.folio,
                    'impreso': enrolamiento.impreso,
                    'fecha_expedicion': enrolamiento.fecha_expedicion
                }
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al marcar empleado como impreso: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class EnrolamientoFamiliarViewSet(AuditoriaUsuarioMixin, viewsets.ModelViewSet):
    queryset = EnrolamientoFamiliar.objects.all().order_by('-id_enrolamiento')
    serializer_class = EnrolamientoFamiliarSerializer

    filter_backends = [filters.SearchFilter]
    search_fields = ['rfc', 'num_empleado', 'nombre', 'paterno', 'folio_familiares']

    @action(detail=False, methods=['get'], url_path='obtener-folio-maximo-familiares')
    def obtener_folio_maximo_familiares(self, request):
        """
        Retorna el folio mÃ¡ximo actual de familiares con consecutivo independiente.
        Usa el campo folio_familiares y mantiene ceros a la izquierda.
        """
        try:
            folios = EnrolamientoFamiliar.objects.exclude(
                Q(folio_familiares__isnull=True) | Q(folio_familiares='')
            ).values_list('folio_familiares', flat=True)

            if not folios:
                return Response({
                    'status': 'success',
                    'folio_maximo': '000000',
                    'siguiente_folio': '000001'
                }, status=status.HTTP_200_OK)

            folio_max_valor = 0
            folio_max_str = ''
            longitud_formato = 6

            for folio in folios:
                folio_str = str(folio).strip()
                solo_digitos = ''.join(ch for ch in folio_str if ch.isdigit())
                if not solo_digitos:
                    continue

                valor = int(solo_digitos)
                if valor > folio_max_valor:
                    folio_max_valor = valor
                    folio_max_str = folio_str
                    longitud_formato = max(len(folio_str), 6)

            if folio_max_valor == 0:
                return Response({
                    'status': 'success',
                    'folio_maximo': '000000',
                    'siguiente_folio': '000001'
                }, status=status.HTTP_200_OK)

            siguiente_valor = folio_max_valor + 1
            siguiente_folio = str(siguiente_valor).zfill(longitud_formato)

            return Response({
                'status': 'success',
                'folio_maximo': folio_max_str,
                'siguiente_folio': siguiente_folio
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al obtener folio mÃ¡ximo de familiares: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='marcar-impreso')
    def marcar_impreso(self, request, pk=None):
        try:
            enrolamiento = self.get_object()

            enrolamiento.impreso = 1
            fecha_expedicion_payload = request.data.get('fecha_expedicion')

            if not enrolamiento.fecha_expedicion:
                enrolamiento.fecha_expedicion = fecha_expedicion_payload or timezone.now().date()
            enrolamiento.save()

            return Response({
                'status': 'success',
                'mensaje': f'Credencial familiar {enrolamiento.folio_familiares or enrolamiento.id_enrolamiento} marcada como impresa',
                'data': {
                    'id_enrolamiento': enrolamiento.id_enrolamiento,
                    'folio_familiares': enrolamiento.folio_familiares,
                    'impreso': enrolamiento.impreso,
                    'fecha_expedicion': enrolamiento.fecha_expedicion
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al marcar familiar como impreso: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SigPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 200


class SigViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SicreTblSig.objects.all().order_by('nombres')
    serializer_class = SigSerializer
    pagination_class = SigPagination

    def get_permissions(self):
        # Unica accion de este viewset que exige sesion: inserta una
        # solicitud que Control_De_Plazas_Backend recoge y ejecuta, asi que
        # no se deja tan abierta como el resto (lectura del roster, todo
        # AllowAny).
        if self.action == 'forzar_actualizacion':
            return [IsAuthenticated()]
        return super().get_permissions()

    filter_backends = [filters.SearchFilter]
    search_fields = [
        'empleado_anam', 'no_empleado', 'curp', 'nombres',
        'primer_apellido', 'segundo_apellido', 'area', 'cargo',
        'estatus', 'estado_hum', 'estado_nom', 'firma_drh', 'cargo_drh',
    ]

    COLUMN_ALIASES = {
        'EMPLEADO_ANAM': ['EMPLEADO_ANAM', 'EMPLEADO ANAM'],
        'NO_EMPLEADO': ['NO_EMPLEADO', 'NO EMPLEADO', 'NUM_EMPLEADO'],
        'CURP': ['CURP'],
        'NOMBRES': ['NOMBRES', 'NOMBRE'],
        'PRIMER_APELLIDO': ['PRIMER_APELLIDO', 'PRIMER APELLIDO', 'PRIMER_APE', 'PATERNO'],
        'SEGUNDO_APELLIDO': ['SEGUNDO_APELLIDO', 'SEGUNDO APELLIDO', 'SEGUNDO_APE', 'MATERNO'],
        'AREA': ['AREA', 'ADSCRIPCION'],
        'CARGO': ['CARGO', 'PUESTO'],
        'FECHA_EXPEDICION': ['FECHA_EXPEDICION', 'FECHA EXPEDICION'],
        'FIRMA_DRH': ['FIRMA_DRH', 'FIRMA DRH'],
        'CARGO_DRH': ['CARGO_DRH', 'CARGO DRH'],
        'QR': ['QR'],
        'ESTATUS': ['ESTATUS', 'STATUS'],
        'ESTADO_HUM': ['ESTADO_HUM', 'ESTADO HUM'],
        'ESTADO_NOM': ['ESTADO_NOM', 'ESTADO NOM'],
    }

    REQUIRED_COLUMNS = list(COLUMN_ALIASES.keys())

    @action(detail=False, methods=['get'], url_path='ultima-actualizacion')
    def ultima_actualizacion(self, request):
        """
        Solo la fecha de sincronizacion mas reciente del roster (un MAX()),
        SIN traer las ~16 mil filas -- para el badge del header, que debe
        conocerla en cuanto se entra al sistema (login), no solo cuando
        alguien visita "Imprimir credenciales" y paga el costo del dataset
        completo (ver `todos`, mismo criterio: el sync de Celery escribe el
        mismo valor en todas las filas de un lote, se toma el maximo por si
        alguna vez quedara una sincronizacion parcial).
        """
        maxima = SicreTblSig.objects.aggregate(maxima=Max('fecha_actualizacion'))['maxima']
        return Response({'status': 'success', 'fecha_actualizacion': maxima})

    @action(detail=False, methods=['post'], url_path='forzar-actualizacion')
    def forzar_actualizacion(self, request):
        """
        Dispara manualmente la sincronizacion del roster (boton "Actualizar"
        del header). Control_De_Plazas_Backend (el proceso de Celery que
        corre la tarea `importar_poblado_credenciales`) vive SIEMPRE en una
        PC Windows local (necesita Edge/Selenium para el SIG) y no expone
        -- ni va a exponer -- un endpoint HTTP: en vez de eso, revisa cada
        30 s la tabla EjeCentral.plantilla_solicitudpobladocredenciales
        (mismo servidor MySQL que sicre_db) y corre la tarea el solo cuando
        encuentra una fila sin atender. Aqui basta con insertar la solicitud.

        El frontend detecta que la tarea ya termino sondeando
        `ultima-actualizacion` con el intervalo acelerado (ver
        RosterSyncService.activarSondeoRapido) hasta que la fecha cambie, no
        por ninguna respuesta de este endpoint.
        """
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO EjeCentral.plantilla_solicitudpobladocredenciales
                        (creado_en, atendido)
                    VALUES (NOW(6), 0)
                    """
                )
        except Exception as exc:
            return Response(
                {
                    'status': 'error',
                    'mensaje': f'No se pudo registrar la solicitud de actualización: {exc}',
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )

        return Response({'status': 'success'})

    @action(detail=False, methods=['get'], url_path='todos')
    def todos(self, request):
        """
        Dataset completo del roster SIG, sin paginar, para filtrado client-side
        en la pantalla "Imprimir credenciales".

        Usa .values() en vez del serializer porque para ~16k registros es
        notablemente mas rapido (evita instanciar un modelo y un serializer por
        fila). El payload ronda los 8 MB sin comprimir, ~1 MB con GZip
        (GZipMiddleware esta activo en settings.py).
        """
        registros = list(
            SicreTblSig.objects.all()
            .order_by('nombres', 'primer_apellido')
            .values()
        )

        # Nombre corto del area, resuelto aqui y no en el front: el catalogo
        # son 66 filas y el cruce es un diccionario en memoria, mucho mas
        # barato que una segunda peticion desde el navegador. Ademas la
        # equivalencia queda en un solo lugar.
        #
        # El cruce va por nombre NORMALIZADO: el roster escribe el area en
        # mayusculas y el catalogo en Title Case, asi que por texto crudo no
        # empata ni una sola fila (comprobado: 0 de 16 431).
        #
        # El catalogo guarda ya resuelto el texto que se imprime (las aduanas
        # traen 'ANAM'), asi que aqui no hay ninguna regla: lo que diga la
        # tabla es lo que sale en la credencial. Eso es lo que permite que se
        # edite desde la pantalla de Catalogos sin tocar codigo.
        catalogo = dict(
            UnidadAdministrativa.objects
            .filter(activo=True)
            .values_list('nombre_normalizado', 'nombre_compactado')
        )

        for registro in registros:
            area = registro.get('area')
            registro['area_compactada'] = (
                catalogo.get(UnidadAdministrativa.normalizar(area)) if area else None
            )

        return Response({
            'status': 'success',
            'total': len(registros),
            'registros': registros,
        })

    @staticmethod
    def _normalize_header(value):
        return str(value or '').strip().upper().replace('.', '').replace('-', '_').replace(' ', '_')

    @staticmethod
    def _clean_text(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _clean_identifier(value, upper=False):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None

        if isinstance(value, float) and value.is_integer():
            text = str(int(value))
        else:
            text = str(value).strip()

        if not text:
            return None

        if text.endswith('.0') and text.replace('.', '', 1).isdigit():
            text = text[:-2]

        return text.upper() if upper else text

    @staticmethod
    def _parse_date(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        parsed = pd.to_datetime(value, errors='coerce')
        if pd.isna(parsed):
            return None
        return parsed.date()

    def _build_dataframe(self, archivo):
        if hasattr(archivo, 'seek'):
            archivo.seek(0)

        df = pd.read_excel(archivo)
        if df.empty:
            return df, []

        normalized_map = {self._normalize_header(col): col for col in df.columns}
        selected_columns = {}
        missing = []

        for canonical, aliases in self.COLUMN_ALIASES.items():
            source_col = None
            for alias in aliases:
                normalized_alias = self._normalize_header(alias)
                if normalized_alias in normalized_map:
                    source_col = normalized_map[normalized_alias]
                    break

            if source_col is None:
                missing.append(canonical)
            else:
                selected_columns[canonical] = source_col

        if missing:
            return None, missing

        normalized_df = pd.DataFrame({
            canonical: df[source]
            for canonical, source in selected_columns.items()
        })

        normalized_df['EMPLEADO_ANAM'] = normalized_df['EMPLEADO_ANAM'].apply(lambda x: self._clean_identifier(x) or '')
        normalized_df['CURP'] = normalized_df['CURP'].apply(lambda x: self._clean_identifier(x, upper=True) or '')

        normalized_df = normalized_df[normalized_df['EMPLEADO_ANAM'] != '']
        normalized_df = normalized_df.drop_duplicates(subset=['EMPLEADO_ANAM'], keep='last')

        return normalized_df, []

    def _sig_payload_from_row(self, row):
        return {
            'empleado_anam': self._clean_identifier(row.get('EMPLEADO_ANAM')),
            'no_empleado': self._clean_identifier(row.get('NO_EMPLEADO')),
            'curp': self._clean_identifier(row.get('CURP'), upper=True),
            'nombres': self._clean_text(row.get('NOMBRES')),
            'primer_apellido': self._clean_text(row.get('PRIMER_APELLIDO')),
            'segundo_apellido': self._clean_text(row.get('SEGUNDO_APELLIDO')),
            'area': self._clean_text(row.get('AREA')),
            'cargo': self._clean_text(row.get('CARGO')),
            'fecha_expedicion': self._parse_date(row.get('FECHA_EXPEDICION')),
            'firma_drh': self._clean_text(row.get('FIRMA_DRH')),
            'cargo_drh': self._clean_text(row.get('CARGO_DRH')),
            'qr': self._clean_text(row.get('QR')),
            'estatus': self._clean_text(row.get('ESTATUS')),
            'estado_hum': self._clean_text(row.get('ESTADO_HUM')),
            'estado_nom': self._clean_text(row.get('ESTADO_NOM')),
        }

    @action(detail=False, methods=['POST'], serializer_class=ArchivoExcelSerializer)
    def previsualizar_excel(self, request):
        serializer = ArchivoExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            archivo = serializer.validated_data['archivo']
            df, missing = self._build_dataframe(archivo)

            if missing:
                return Response({
                    'status': 'error',
                    'mensaje': 'El archivo no contiene todas las columnas requeridas de SIG.',
                    'columnas_faltantes': missing
                }, status=status.HTTP_400_BAD_REQUEST)

            total_registros = 0 if df is None else len(df)
            preview_limit = min(max(int(request.query_params.get('preview_limit', 200)), 1), 500)
            preview_df = df.head(preview_limit) if df is not None else pd.DataFrame()

            registros_preview = []
            for _, row in preview_df.iterrows():
                registro = self._sig_payload_from_row(row)
                registros_preview.append(registro)

            return Response({
                'status': 'success',
                'mensaje': 'Vista previa generada correctamente',
                'total_registros': total_registros,
                'registros': registros_preview,
                'preview_limit': preview_limit,
                'preview_registros': len(registros_preview),
                'preview_parcial': total_registros > len(registros_preview)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al leer el archivo: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'], serializer_class=ArchivoExcelSerializer)
    def subir_excel(self, request):
        serializer = ArchivoExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            archivo = serializer.validated_data['archivo']
            df, missing = self._build_dataframe(archivo)

            if missing:
                return Response({
                    'status': 'error',
                    'mensaje': 'El archivo no contiene todas las columnas requeridas de SIG.',
                    'columnas_faltantes': missing
                }, status=status.HTTP_400_BAD_REQUEST)

            if df is None or df.empty:
                return Response({
                    'status': 'error',
                    'mensaje': 'No hay registros vÃ¡lidos para procesar.'
                }, status=status.HTTP_400_BAD_REQUEST)

            rows = [self._sig_payload_from_row(row) for _, row in df.iterrows()]

            # --- ValidaciÃ³n de duplicados en el archivo ---
            curp_count = {}
            num_emp_count = {}
            for r in rows:
                c = (r.get('curp') or '').strip()
                n = (r.get('no_empleado') or '').strip()
                if c:
                    curp_count[c] = curp_count.get(c, 0) + 1
                if n:
                    num_emp_count[n] = num_emp_count.get(n, 0) + 1

            curps_dup = [c for c, cnt in curp_count.items() if cnt > 1]
            nums_dup  = [n for n, cnt in num_emp_count.items() if cnt > 1]

            if curps_dup or nums_dup:
                errores = {}
                if curps_dup:
                    errores['curps_duplicadas'] = curps_dup
                if nums_dup:
                    errores['numeros_empleado_duplicados'] = nums_dup
                return Response({
                    'status': 'error',
                    'mensaje': 'El archivo contiene registros duplicados. No puede haber mÃ¡s de un empleado con la misma CURP o el mismo nÃºmero de empleado.',
                    **errores
                }, status=status.HTTP_400_BAD_REQUEST)
            # --- Fin validaciÃ³n duplicados ---

            rows_by_empleado = {}
            for row in rows:
                empleado_key = row.get('empleado_anam')
                if not empleado_key:
                    continue
                rows_by_empleado[empleado_key] = row

            rows = list(rows_by_empleado.values())
            empleados_anam = [row['empleado_anam'] for row in rows if row.get('empleado_anam')]
            curps = [row['curp'] for row in rows if row.get('curp')]

            sig_create = []
            sig_update = []
            enr_create = []
            enr_update = []

            sig_existing = SicreTblSig.objects.in_bulk(empleados_anam)
            sig_existing_by_curp_qs = SicreTblSig.objects.filter(curp__in=curps)
            sig_existing_by_curp = {}
            for obj in sig_existing_by_curp_qs:
                if obj.curp and obj.curp not in sig_existing_by_curp:
                    sig_existing_by_curp[obj.curp] = obj

            enr_existing_qs = Enrolamiento.objects.filter(curp__in=curps).order_by('-id_enrolamiento')
            enr_existing = {}
            for obj in enr_existing_qs:
                if obj.curp and obj.curp not in enr_existing:
                    enr_existing[obj.curp] = obj

            max_id = SicreTblSig.objects.aggregate(max_id=models.Max('id')).get('max_id') or 0
            next_id = int(max_id)
            now_ts = timezone.now()

            def normalized_text(value):
                if value is None:
                    return ''
                return str(value).strip()

            def same_text(current, incoming):
                return normalized_text(current) == normalized_text(incoming)

            def same_date(current, incoming):
                return current == incoming

            for row in rows:
                empleado_anam = row['empleado_anam']
                curp = row.get('curp')
                if not empleado_anam:
                    continue

                sig_obj = sig_existing.get(empleado_anam)
                if not sig_obj and curp:
                    sig_obj = sig_existing_by_curp.get(curp)
                if sig_obj:
                    sig_changed = False

                    if sig_obj.id is None:
                        next_id += 1
                        sig_obj.id = next_id
                        sig_changed = True

                    sig_new_values = {
                        'no_empleado': row['no_empleado'],
                        'curp': row['curp'],
                        'nombres': row['nombres'],
                        'primer_apellido': row['primer_apellido'],
                        'segundo_apellido': row['segundo_apellido'],
                        'area': row['area'],
                        'cargo': row['cargo'],
                        'fecha_expedicion': row['fecha_expedicion'],
                        'firma_drh': row['firma_drh'],
                        'cargo_drh': row['cargo_drh'],
                        'qr': row['qr'],
                        'estatus': row['estatus'],
                        'estado_hum': row['estado_hum'],
                        'estado_nom': row['estado_nom'],
                    }

                    for field_name, new_value in sig_new_values.items():
                        current_value = getattr(sig_obj, field_name)
                        is_same = same_date(current_value, new_value) if field_name == 'fecha_expedicion' else same_text(current_value, new_value)
                        if not is_same:
                            setattr(sig_obj, field_name, new_value)
                            sig_changed = True

                    if sig_changed:
                        sig_obj.fecha_actualizacion = now_ts
                        sig_update.append(sig_obj)
                else:
                    next_id += 1
                    sig_create.append(SicreTblSig(
                        id=next_id,
                        fecha_actualizacion=now_ts,
                        **row
                    ))

                nombre = (row.get('nombres') or '').strip()
                paterno = (row.get('primer_apellido') or '').strip()
                materno = (row.get('segundo_apellido') or '').strip()
                apellidos = f"{paterno} {materno}".strip()
                estatus = (row.get('estatus') or '').lower()
                activo_flag = 0 if 'baja' in estatus else 1
                rfc_ref = (row.get('empleado_anam') or row.get('no_empleado') or curp or '').strip()

                if not curp:
                    continue

                enr_obj = enr_existing.get(curp)
                if enr_obj:
                    enr_changed = False
                    enr_new_values = {
                        'rfc': rfc_ref,
                        'num_empleado': row.get('no_empleado'),
                        'nombre': nombre,
                        'paterno': paterno,
                        'materno': materno,
                        'apellidos': apellidos,
                        'puesto': row.get('cargo'),
                        'adscripcion': row.get('area'),
                        'activo': activo_flag,
                        'nivel_credencial': _cargo_a_nivel(row.get('cargo') or ''),
                        'layout_credencial': 'ANAM_2025',
                    }

                    for field_name, new_value in enr_new_values.items():
                        current_value = getattr(enr_obj, field_name)
                        is_same = current_value == new_value if field_name == 'activo' else same_text(current_value, new_value)
                        if not is_same:
                            setattr(enr_obj, field_name, new_value)
                            enr_changed = True

                    if enr_changed:
                        enr_obj.fecha_modificacion = timezone.now()
                        enr_update.append(enr_obj)
                else:
                    enr_create.append(Enrolamiento(
                        rfc=rfc_ref,
                        num_empleado=row.get('no_empleado'),
                        curp=curp,
                        nombre=nombre,
                        paterno=paterno,
                        materno=materno,
                        apellidos=apellidos,
                        puesto=row.get('cargo'),
                        adscripcion=row.get('area'),
                        activo=activo_flag,
                        fecha_enrolamiento=timezone.now(),
                        nivel_credencial=_cargo_a_nivel(row.get('cargo') or ''),
                        layout_credencial='ANAM_2025'
                    ))

            sig_update = list({obj.empleado_anam: obj for obj in sig_update}.values())
            enr_update = list({obj.id_enrolamiento: obj for obj in enr_update}.values())

            if not sig_create and not sig_update and not enr_create and not enr_update:
                resumen_sin_cambios = {
                    'registros_archivo': len(rows),
                    'sig_creados': 0,
                    'sig_actualizados': 0,
                    'enrolamiento_creados': 0,
                    'enrolamiento_actualizados': 0
                }
                return Response({
                    'status': 'success',
                    'mensaje': 'No se encontraron nuevos empleados ni actualizaciones de registros.',
                    'resumen': resumen_sin_cambios
                }, status=status.HTTP_200_OK)

            with transaction.atomic():
                if sig_create:
                    SicreTblSig.objects.bulk_create(sig_create, batch_size=1000)
                if sig_update:
                    SicreTblSig.objects.bulk_update(
                        sig_update,
                        ['id', 'no_empleado', 'curp', 'nombres', 'primer_apellido', 'segundo_apellido', 'area', 'cargo', 'fecha_expedicion', 'firma_drh', 'cargo_drh', 'qr', 'estatus', 'estado_hum', 'estado_nom', 'fecha_actualizacion'],
                        batch_size=1000
                    )

                if enr_create:
                    Enrolamiento.objects.bulk_create(enr_create, batch_size=1000)
                if enr_update:
                    Enrolamiento.objects.bulk_update(
                        enr_update,
                        ['rfc', 'num_empleado', 'nombre', 'paterno', 'materno', 'apellidos', 'puesto', 'adscripcion', 'activo', 'nivel_credencial', 'layout_credencial', 'fecha_modificacion'],
                        batch_size=1000
                    )

            resumen = {
                'registros_archivo': len(rows),
                'sig_creados': len(sig_create),
                'sig_actualizados': len(sig_update),
                'enrolamiento_creados': len(enr_create),
                'enrolamiento_actualizados': len(enr_update),
                'detalle_empleados': [
                    {
                        'no_empleado': obj.no_empleado or '',
                        'nombre': f"{obj.nombres or ''} {obj.primer_apellido or ''} {obj.segundo_apellido or ''}".strip(),
                        'tipo': 'Nuevo ingreso'
                    }
                    for obj in sig_create
                ] + [
                    {
                        'no_empleado': obj.no_empleado or '',
                        'nombre': f"{obj.nombres or ''} {obj.primer_apellido or ''} {obj.segundo_apellido or ''}".strip(),
                        'tipo': 'ActualizaciÃ³n'
                    }
                    for obj in sig_update
                ]
            }

            sin_cambios = (
                resumen['sig_creados'] == 0 and
                resumen['sig_actualizados'] == 0 and
                resumen['enrolamiento_creados'] == 0 and
                resumen['enrolamiento_actualizados'] == 0
            )

            return Response({
                'status': 'success',
                'mensaje': 'No se encontraron nuevos empleados ni actualizaciones de registros.' if sin_cambios else 'Consulta de SIG actualizada exitosamente.',
                'resumen': resumen
            }, status=status.HTTP_200_OK if sin_cambios else status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error procesando el archivo: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='historial-cargas')
    def historial_cargas(self, request):
        try:
            resumen_estatus = list(
                SicreTblSig.objects.values('estatus').annotate(total=Count('empleado_anam')).order_by('-total')[:10]
            )
            preview = list(
                SicreTblSig.objects.values('empleado_anam', 'curp', 'nombres', 'primer_apellido', 'segundo_apellido', 'fecha_actualizacion')[:10]
            )

            return Response({
                'status': 'success',
                'total_registros': SicreTblSig.objects.count(),
                'por_estatus': resumen_estatus,
                'preview': preview
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'mensaje': f'Error al obtener historial: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='detalle-lote/(?P<lote_id>[0-9]+)')
    def detalle_lote(self, request, lote_id=None):
        return Response({
            'status': 'error',
            'mensaje': 'Detalle por lote ya no aplica en el nuevo esquema SIG.'
        }, status=status.HTTP_410_GONE)


class CustomLoginView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            identifier = serializer.validated_data['email']
            password = serializer.validated_data['password']

            # Intenta autenticar por username directo
            user = authenticate(request, username=identifier, password=password)

            # Si falla, busca el username asociado al email ingresado
            if user is None and '@' in identifier:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                try:
                    db_user = User.objects.get(email__iexact=identifier)
                    user = authenticate(request, username=db_user.username, password=password)
                except User.DoesNotExist:
                    pass

            if user is not None:
                if not user.is_active:
                    return Response({
                        'status': 403,
                        'message': 'Usuario inactivo'
                    }, status=status.HTTP_200_OK)

                token, created = Token.objects.get_or_create(user=user)

                num_empleado = getattr(getattr(user, 'perfil', None), 'num_empleado', '') or ''
                foto = None
                if num_empleado:
                    ruta_foto = media_utils.resolver_foto(num_empleado)
                    foto = media_utils.url_publica(ruta_foto) if ruta_foto else None

                return Response({
                    'status': 200,
                    'message': 'Login exitoso',
                    'model': {
                        'token': token.key,
                        # `tokenWs` es el nombre que en realidad lee
                        # TokenInterceptor -- se manda tambien asi para no
                        # tener que tocar el interceptor (corre en cada
                        # peticion del sistema). Antes de este cambio ningun
                        # login real mandaba Authorization en absoluto: no
                        # importaba porque el backend no autenticaba a nadie
                        # (ver CLAUDE.md, gotcha #3), pero ahora que las
                        # pantallas de administracion SI exigen sesion, hacia
                        # falta que coincidiera.
                        'tokenWs': token.key,
                        'idUsuario': user.id,
                        'username': user.username,
                        'nombreCompleto': f"{user.first_name} {user.last_name}",
                        'email': user.email,
                        'unidadAdscripcion': getattr(user, 'adscripcion', ''),
                        'area': 'SICRE',
                        'numEmpleado': num_empleado,
                        'foto': foto,
                        # Se conserva por compatibilidad -- nada del sistema
                        # nuevo de permisos lo usa ya (ver `permisos` /
                        # `esSuperusuario`), pero quitarlo de golpe podria
                        # romper algo que todavia lo lea.
                        'idUsuarioRol': 9999 if user.is_superuser else (2 if user.is_staff else 4),
                        'esSuperusuario': user.is_superuser,
                        # Codenames sin el prefijo de la app ('ver_plantillas',
                        # no 'credencializacion.ver_plantillas'): son unicos
                        # dentro del catalogo y asi el frontend no tiene que
                        # conocer el nombre de la app Django.
                        #
                        # Se filtra contra CODENAMES_CATALOGO_PERMISOS, no solo
                        # contra el prefijo de la app: sin esto, un superusuario
                        # (que Django considera dueno de TODOS los permisos)
                        # recibiria tambien los add/change/delete/view que se
                        # crean automaticamente para cada modelo -- ruido que
                        # no significa nada para este sistema de menus.
                        'permisos': sorted({
                            p.split('.', 1)[1] for p in user.get_all_permissions()
                            if p.startswith('credencializacion.')
                            and p.split('.', 1)[1] in dict(CODENAMES_CATALOGO_PERMISOS)
                        }),
                    }
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'status': 401,
                    'message': 'Credenciales incorrectas'
                }, status=status.HTTP_200_OK) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def foto_firma_empleado(request, emplid):
    """
    Obtiene foto y firma de safirho_db.NW_EMPL_FOTO_ANAM para un EMPLID dado.
    Prueba con el valor tal cual, zero-padded a 11 dÃ­gitos y sin ceros a la izquierda.
    """
    emplid_str = str(emplid).strip()
    emplid_padded = emplid_str.zfill(11)
    emplid_stripped = emplid_str.lstrip('0') or '0'

    variantes = list({emplid_str, emplid_padded, emplid_stripped})
    placeholders = ','.join(['%s'] * len(variantes))

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT FOTO, FIRMA, TipoMime FROM safirho_db.NW_EMPL_FOTO_ANAM WHERE EMPLID IN ({placeholders}) LIMIT 1",
                variantes
            )
            row = cursor.fetchone()

        if not row:
            return Response({'foto': None, 'firma': None, 'encontrado': False})

        foto_raw, firma_raw, tipo_mime = row
        mime = tipo_mime or 'image/jpeg'

        def blob_to_b64(raw):
            if raw is None:
                return None
            if isinstance(raw, memoryview):
                raw = bytes(raw)
            elif isinstance(raw, bytearray):
                raw = bytes(raw)
            elif not isinstance(raw, bytes):
                raw = bytes(raw)
            if not raw:
                return None
            return f'data:{mime};base64,' + base64.b64encode(raw).decode('utf-8')

        return Response({
            'foto': blob_to_b64(foto_raw),
            'firma': blob_to_b64(firma_raw),
            'encontrado': True,
        })

    except Exception as e:
        return Response({'foto': None, 'firma': None, 'encontrado': False, 'error': str(e)}, status=200)
class CargaMasivaPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 500


class CargaMasivaViewSet(AuditoriaUsuarioMixin, viewsets.ModelViewSet):
    serializer_class = CargaMasivaSerializer
    pagination_class = CargaMasivaPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['rfc', 'curp', 'nombres', 'primer_apellido', 'segundo_apellido',
                     'empleado_anam', 'no_empleado', 'area', 'cargo', 'estatus']

    def get_queryset(self):
        qs = CargaMasiva.objects.filter(activo=True).order_by('nombres', 'primer_apellido')
        lote = self.request.query_params.get('lote')
        if lote:
            qs = qs.filter(lote=lote)
        return qs

    @action(detail=False, methods=['post'], url_path='auto-guardado')
    def auto_guardado(self, request):
        lote = request.data.get('lote')
        rfc = request.data.get('rfc')
        record_id = request.data.get('id')

        if not lote or not rfc:
            return Response({'error': 'Lote y RFC son requeridos.'}, status=status.HTTP_400_BAD_REQUEST)

        instance = None
        if record_id:
            instance = CargaMasiva.objects.filter(id=record_id, activo=True).first()
        if not instance:
            instance = CargaMasiva.objects.filter(lote=lote, rfc=rfc, activo=True).first()

        if instance:
            serializer = self.get_serializer(instance, data=request.data, partial=True)
        else:
            serializer = self.get_serializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK if instance else status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='siguiente-lote')
    def siguiente_lote(self, request):
        ultimo = CargaMasiva.objects.exclude(lote__isnull=True).order_by('-id').first()
        siguiente_num = 1
        if ultimo and ultimo.lote:
            import re
            match = re.search(r'LOTE-(\d+)', ultimo.lote)
            if match:
                siguiente_num = int(match.group(1)) + 1
        return Response({'lote': f'LOTE-{siguiente_num:05d}'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='cancelar-lote')
    def cancelar_lote(self, request):
        lote = request.data.get('lote')
        if not lote:
            return Response({'error': 'Lote es requerido.'}, status=status.HTTP_400_BAD_REQUEST)
        
        registros = CargaMasiva.objects.filter(lote=lote, activo=True)
        count = registros.count()
        registros.update(activo=False)
        
        return Response({'message': f'Lote {lote} cancelado.', 'registros_eliminados': count}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='cargar-lote-excel')
    def cargar_lote_excel(self, request):
        """
        Recibe un lote existente y un archivo Excel con columnas de SIG.
        Para cada fila del Excel busca en sicre_tbl_carga_masiva el registro
        cuyo campo rfc coincida con la columna CURP del Excel (dentro del lote).
        Si encuentra coincidencia actualiza todos los campos del registro.
        Si no encuentra, lo reporta en 'no_encontrados'.

        Respuesta:
        {
          "lote": "LOTE-00001",
          "total": 50,
          "actualizados": 45,
          "no_encontrados": [ { "curp": "...", "nombre": "..." }, ... ]
        }
        """
        archivo = request.FILES.get('archivo')
        lote = request.data.get('lote')

        if not archivo:
            return Response({'error': 'Se requiere el archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        if not lote:
            return Response({'error': 'Se requiere seleccionar un lote.'}, status=status.HTTP_400_BAD_REQUEST)

        if not CargaMasiva.objects.filter(lote=lote, activo=True).exists():
            return Response({'error': f'El lote {lote} no existe o no tiene registros activos.'}, status=status.HTTP_404_NOT_FOUND)

        # ── 1. Parsear Excel ───────────────────────────────────────────────────
        sig_vs = SigViewSet()
        df, missing = sig_vs._build_dataframe(archivo)

        if missing:
            return Response({
                'error': 'El archivo no contiene todas las columnas requeridas.',
                'columnas_faltantes': missing
            }, status=status.HTTP_400_BAD_REQUEST)

        if df is None or df.empty:
            return Response({'error': 'El archivo no tiene registros válidos.'}, status=status.HTTP_400_BAD_REQUEST)

        rows = [sig_vs._sig_payload_from_row(row) for _, row in df.iterrows()]

        # ── 2. Actualizar registros del lote haciendo match CURP → rfc ─────────
        campos_sig = [
            'empleado_anam', 'no_empleado', 'curp', 'nombres',
            'primer_apellido', 'segundo_apellido', 'area', 'cargo',
            'fecha_expedicion', 'firma_drh', 'cargo_drh', 'qr',
            'estatus', 'estado_hum', 'estado_nom',
        ]

        actualizados = 0
        no_encontrados = []

        with transaction.atomic():
            for row in rows:
                curp = (row.get('curp') or '').strip()
                nombre = ' '.join(filter(None, [
                    row.get('nombres'), row.get('primer_apellido'), row.get('segundo_apellido')
                ])).strip()

                try:
                    registro = CargaMasiva.objects.get(lote=lote, rfc=curp, activo=True)
                    for campo in campos_sig:
                        valor = row.get(campo)
                        if valor is not None:
                            setattr(registro, campo, valor)
                    registro.save(update_fields=campos_sig)
                    actualizados += 1
                except CargaMasiva.DoesNotExist:
                    no_encontrados.append({'curp': curp, 'nombre': nombre})
                except CargaMasiva.MultipleObjectsReturned:
                    # Si hay duplicados en el lote, actualiza el primero
                    registro = CargaMasiva.objects.filter(lote=lote, rfc=curp, activo=True).first()
                    for campo in campos_sig:
                        valor = row.get(campo)
                        if valor is not None:
                            setattr(registro, campo, valor)
                    registro.save(update_fields=campos_sig)
                    actualizados += 1

        return Response({
            'lote': lote,
            'total': len(rows),
            'actualizados': actualizados,
            'no_encontrados': no_encontrados,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='lotes-resumen')
    def lotes_resumen(self, request):
        from django.db.models import Count, Max, Min, Q as DQ
        lotes = (
            CargaMasiva.objects
            .filter(activo=True)
            .values('lote')
            .annotate(
                total=Count('id'),
                total_fotos=Count('id', filter=DQ(foto__isnull=False)),
                total_firmas=Count('id', filter=DQ(firma__isnull=False)),
                fecha_inicio=Min('fecha_enrolamiento'),
                fecha_ultimo=Max('fecha_enrolamiento'),
            )
            .order_by('-fecha_ultimo')
        )
        return Response(list(lotes), status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='progreso-lote')
    def progreso_lote(self, request):
        lote = request.query_params.get('lote')
        sin_imagenes = request.query_params.get('sin_imagenes', '0') == '1'
        if not lote:
            return Response({'error': 'Lote es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        registros_qs = list(CargaMasiva.objects.filter(lote=lote, activo=True).order_by('id'))
        total_enrolados = len(registros_qs)
        total_fotos = sum(1 for r in registros_qs if r.foto)
        total_firmas = sum(1 for r in registros_qs if r.firma)

        if sin_imagenes:
            datos_registros = [
                {
                    'id': r.id,
                    'rfc': r.rfc,
                    'nombre': r.nombre,
                    'nombres': r.nombres,
                    'primer_apellido': r.primer_apellido,
                    'segundo_apellido': r.segundo_apellido,
                    'no_empleado': r.no_empleado,
                    'empleado_anam': r.empleado_anam,
                    'curp': r.curp,
                    'area': r.area,
                    'cargo': r.cargo,
                    'lote': r.lote,
                    'fecha_enrolamiento': r.fecha_enrolamiento.isoformat() if r.fecha_enrolamiento else None,
                    'has_foto': bool(r.foto),
                    'has_firma': bool(r.firma),
                    'activo': r.activo,
                    'nivel_credencial': r.nivel_credencial,
                    'layout_credencial': r.layout_credencial,
                    'nuevo_laredo': r.nuevo_laredo,
                    'folio': r.folio,
                    'fecha_expedicion': r.fecha_expedicion.isoformat() if r.fecha_expedicion else None,
                    'fin_vig': r.fin_vig.isoformat() if r.fin_vig else None,
                    'inicio_vig': r.inicio_vig.isoformat() if r.inicio_vig else None,
                }
                for r in registros_qs
            ]
        else:
            datos_registros = self.get_serializer(registros_qs, many=True).data

        return Response({
            'lote': lote,
            'total_enrolados': total_enrolados,
            'total_fotografias': total_fotos,
            'total_firmas': total_firmas,
            'registros': datos_registros,
        }, status=status.HTTP_200_OK)


# ==========================================================================
# NUEVO ESQUEMA: enrolamientos con medios en disco + plantillas tipo canvas
# ==========================================================================

class EnrolamientoCredencialPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 500


def _congelar_snapshot(canvas_json, medios=None):
    """
    Deja un lienzo de Fabric listo para archivarse como constancia de lo
    impreso: cada imagen que apunta a MEDIA_ROOT se copia al archivo historico
    y su `src` se reescribe a esa copia inmutable.

    Sin esto el snapshot seria solo *aparentemente* fiel. Las imagenes viajan
    como URL (`http://host/media/fotos/123.png?v=1786056998`), no como bytes,
    de modo que apuntan al archivo VIVO: si manana recapturan la foto de esa
    persona, la credencial impresa hace seis meses empezaria a mostrar la cara
    nueva sobre el folio viejo. Peor todavia, el `?v=<mtime>` cambia al
    reemplazarla y la imagen dejaria de cargar del todo.

    Se recorren tambien los grupos anidados y el `backgroundImage` -- el fondo
    de la plantilla vive ahi, y las plantillas son editables, asi que tampoco
    se puede confiar en que siga siendo el mismo.

    Los data-URI (el QR, que se genera en el navegador) se dejan intactos: ya
    son bytes, no referencias.

    Si se pasa un diccionario en `medios`, se anota ahi que archivo quedo como
    foto, firma y fondo. El rol solo se puede saber AQUI, por la carpeta de
    origen: una vez archivado, el nombre es un hash y ya no dice nada.
    """
    if not isinstance(canvas_json, dict):
        return None

    carpetas = {
        media_utils.carpeta_fotos(): 'foto',
        media_utils.carpeta_firmas(): 'firma',
        media_utils.carpeta_plantillas(): 'fondo',
    }

    def archivar_nodo(nodo):
        if not isinstance(nodo, dict):
            return

        src = nodo.get('src')
        if isinstance(src, str) and src:
            relativa = media_utils.ruta_relativa_desde_url(src)
            archivada = media_utils.archivar_medio(relativa) if relativa else None
            if archivada:
                # Sin `?v=`: la ruta ya identifica el contenido por su hash,
                # asi que nunca cambia y puede cachearse para siempre.
                nodo['src'] = media_utils.url_publica(archivada, versionada=False)

                if medios is not None:
                    rol = carpetas.get(str(relativa).split('/', 1)[0])
                    # `setdefault`: si la misma cara trae dos imagenes de la
                    # misma carpeta, se conserva la primera en vez de que la
                    # ultima pise a la anterior.
                    if rol:
                        medios.setdefault(rol, archivada)

        for hijo in (nodo.get('objects') or []):
            archivar_nodo(hijo)

    archivar_nodo(canvas_json.get('backgroundImage'))
    for objeto in (canvas_json.get('objects') or []):
        archivar_nodo(objeto)

    return canvas_json


class EnrolamientoCredencialViewSet(AuditoriaUsuarioMixin, viewsets.ModelViewSet):
    """
    Registro de qué pasó al expedir cada credencial (folio, vigencias,
    plantilla usada, estatus de impresión) -- NO los datos del empleado, que
    viven en sicre_tbl_sig y se consultan cruzando por num_empleado. Ver
    docstring del modelo EnrolamientoCredencial.
    """
    # `defer` sobre los lienzos NO es una optimizacion opcional: cada fila
    # lleva ~47 KB de JSON y MySQL mete la fila COMPLETA en el buffer de
    # ordenamiento. Sin esto, cualquier listado ordenado revienta con
    # "Out of sort memory, consider increasing server sort buffer size" en
    # cuanto la tabla crece -- reproducido con apenas unas decenas de filas.
    # Quien necesite el lienzo lo pide por `auditoria-detalle`.
    queryset = (
        EnrolamientoCredencial.objects
        .defer('canvas_frente', 'canvas_reverso')
        .order_by('-id_enrolamiento')
    )
    serializer_class = EnrolamientoCredencialSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['num_empleado', 'folio']
    pagination_class = EnrolamientoCredencialPagination

    # Acciones EXCLUSIVAS de "Inventario de medios" (renombrar un RFC mal
    # tecleado, cruzarlo al num_empleado definitivo, borrar un pendiente):
    # exigen `medios_administrar`. Deliberadamente NO se incluye aqui
    # `guardar-medios-empleado`: la usan tambien "Imprimir credenciales" y la
    # captura normal de "Enrolamiento previo" para su propio flujo del dia a
    # dia (tomar/reemplazar una foto antes de imprimir), que no es una
    # operacion "administrativa" y no deberia depender de ese permiso.
    _ACCIONES_MEDIOS_ADMINISTRAR = {
        'migrar_medios', 'migrar_medios_lote', 'renombrar_medios', 'borrar_medios_previo',
    }
    # Descarga masiva de /media -- permiso propio y distinto de
    # `medios_administrar`: bajar un respaldo completo no es lo mismo que
    # poder renombrar/borrar capturas individuales, y se quiere poder
    # asignar por separado desde /administracion.
    _ACCIONES_MEDIOS_RESPALDO = {'respaldo_fotos', 'respaldo_firmas'}

    def get_permissions(self):
        if self.action in self._ACCIONES_MEDIOS_ADMINISTRAR:
            return [IsAuthenticated(), tiene_permiso('medios_administrar')()]
        if self.action in self._ACCIONES_MEDIOS_RESPALDO:
            return [IsAuthenticated(), tiene_permiso('medios_respaldo')()]
        return super().get_permissions()

    def _respaldo_carpeta(self, subcarpeta, etiqueta):
        """
        Zippea MEDIA_ROOT/<subcarpeta> completa y la manda como descarga.

        Se arma en un archivo temporal SIN NOMBRE (tempfile.TemporaryFile):
        con ~14 mil fotos el zip pesa cientos de MB, y guardarlo en memoria
        (BytesIO) reventaria el proceso. El archivo se borra solo del disco
        en cuanto FileResponse termina de mandarlo y lo cierra.
        """
        carpeta = Path(settings.MEDIA_ROOT) / subcarpeta
        tmp = tempfile.TemporaryFile(suffix='.zip')
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            if carpeta.is_dir():
                for archivo in sorted(carpeta.iterdir()):
                    if archivo.is_file():
                        zf.write(archivo, arcname=archivo.name)
        tamano = tmp.tell()
        tmp.seek(0)

        nombre = f'{etiqueta}_{timezone.now():%Y%m%d_%H%M}.zip'
        respuesta = FileResponse(tmp, as_attachment=True, filename=nombre)
        respuesta['Content-Length'] = tamano
        return respuesta

    @action(detail=False, methods=['get'], url_path='respaldo-fotos')
    def respaldo_fotos(self, request):
        return self._respaldo_carpeta(settings.MEDIA_DIR_FOTOS, 'fotos')

    @action(detail=False, methods=['get'], url_path='respaldo-firmas')
    def respaldo_firmas(self, request):
        return self._respaldo_carpeta(settings.MEDIA_DIR_FIRMAS, 'firmas')

    def get_queryset(self):
        qs = super().get_queryset()

        num_empleado = self.request.query_params.get('num_empleado')
        if num_empleado:
            qs = qs.filter(num_empleado=str(num_empleado).strip())

        solo_activos = self.request.query_params.get('activo')
        if solo_activos is not None and str(solo_activos) != '':
            qs = qs.filter(activo=int(solo_activos))

        return qs

    def perform_update(self, serializer):
        serializer.save(fecha_modificacion=timezone.now())

    @action(detail=False, methods=['get'], url_path='buscar-empleado')
    def buscar_empleado(self, request):
        """
        Busca un empleado por num_empleado en la tabla nueva y, si no existe ahi,
        cae a las tablas historicas (Enrolamiento / SIG) para prellenar la captura.

        Siempre resuelve las rutas de foto/firma existentes en MEDIA_ROOT, de modo
        que se pueda reconstruir la credencial de cualquier empleado ya enrolado.
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registro = EnrolamientoCredencial.objects.filter(num_empleado=num_empleado).first()
        if registro:
            return Response({
                'status': 'success',
                'origen': 'enrolamiento_credencial',
                'datos': EnrolamientoCredencialSerializer(registro).data,
            })

        datos = None
        origen = None

        historico = Enrolamiento.objects.filter(num_empleado=num_empleado).first()
        if historico:
            origen = 'enrolamiento'
            datos = {
                'num_empleado': historico.num_empleado,
                'rfc': historico.rfc,
                'curp': historico.curp,
                'nombre': historico.nombre,
                'paterno': historico.paterno,
                'materno': historico.materno,
                'apellidos': historico.apellidos,
                'puesto': historico.puesto,
                'adscripcion': historico.adscripcion,
                'inicio_vig': historico.inicio_vig,
                'fin_vig': historico.fin_vig,
                'folio': historico.folio,
                'fecha_expedicion': historico.fecha_expedicion,
                'layout_credencial': historico.layout_credencial,
            }
        else:
            sig = SicreTblSig.objects.filter(
                Q(no_empleado=num_empleado) | Q(empleado_anam=num_empleado)
            ).first()
            if sig:
                origen = 'sig'
                datos = {
                    'num_empleado': sig.no_empleado or sig.empleado_anam,
                    'rfc': sig.curp or '',
                    'curp': sig.curp,
                    'nombre': sig.nombres,
                    'paterno': sig.primer_apellido,
                    'materno': sig.segundo_apellido,
                    'apellidos': f"{sig.primer_apellido or ''} {sig.segundo_apellido or ''}".strip(),
                    'puesto': sig.cargo,
                    # Se imprime el nombre CORTO del area (catalogo
                    # sicre_cat_unidad_compactada); el oficial completo queda
                    # en area_completa por si alguna plantilla lo necesita.
                    'area': UnidadAdministrativa.compactar(sig.area),
                    'adscripcion': UnidadAdministrativa.compactar(sig.area),
                    'area_completa': sig.area,
                    'fecha_expedicion': sig.fecha_expedicion,
                }

        if not datos:
            return Response(
                {'status': 'not_found', 'mensaje': f'No se encontro al empleado {num_empleado}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Vincular medios historicos aunque el registro aun no exista en la tabla nueva.
        ruta_foto = media_utils.resolver_foto(num_empleado)
        ruta_firma = media_utils.resolver_firma(num_empleado)
        datos['foto'] = media_utils.url_publica(ruta_foto)
        datos['firma'] = media_utils.url_publica(ruta_firma)
        datos['foto_ruta'] = ruta_foto
        datos['firma_ruta'] = ruta_firma

        return Response({'status': 'success', 'origen': origen, 'datos': datos})

    @action(detail=False, methods=['post'], url_path='registrar-impresion')
    def registrar_impresion(self, request):
        """
        Deja constancia de una credencial impresa. UNA FILA POR IMPRESION.

        Solo lo llama "Imprimir credenciales" despues de generar el PDF, no al
        seleccionar un empleado: de otro modo la tabla acumularia una fila por
        cada uno de los 16 000 registros que alguien abra, y dejaria de ser un
        historial para volverse una copia del roster.

        Los datos personales del roster NO se copian como columnas: viven en
        sicre_tbl_sig y se cruzan por num_empleado. Lo que si queda congelado
        es el LIENZO impreso (ver `_congelar_snapshot`), que es lo unico que
        permite reproducir despues la credencial exacta que se expidio ese dia.
        """
        num_empleado = (request.data.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Snapshot de AMBAS caras, siempre. Sus imagenes quedan archivadas en
        # media/historico/ para que nadie las pueda reemplazar despues, y de
        # paso se anota cual quedo como foto, firma y fondo.
        medios = {}
        canvas_frente = _congelar_snapshot(request.data.get('canvas_frente'), medios)
        canvas_reverso = _congelar_snapshot(request.data.get('canvas_reverso'), medios)

        registro = EnrolamientoCredencial.objects.create(
            num_empleado=num_empleado,
            folio=(request.data.get('folio') or None),
            fecha_expedicion=request.data.get('fecha_expedicion') or timezone.now().date(),
            fin_vig=request.data.get('fin_vig') or None,
            plantilla_credencial=(request.data.get('plantilla_credencial') or None),
            canvas_frente=canvas_frente,
            canvas_reverso=canvas_reverso,
            medios=(medios or None),
            con_ajustes=bool(request.data.get('con_ajustes')),
            activo=1,
            id_usuario_registra=_usuario_actual_id(request),
        )

        return Response(
            {'status': 'success', 'id_enrolamiento': registro.id_enrolamiento},
            status=status.HTTP_201_CREATED,
        )

    # ----------------------------------------------------------------
    # Auditoria de credenciales impresas
    # ----------------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='auditoria')
    def auditoria(self, request):
        """
        Historial completo de credenciales impresas, para la pantalla de
        auditoria. Una fila por impresion.

        NO trae los lienzos: pesan ~47 KB cada uno y la pantalla solo los
        necesita al abrir una credencial concreta (ver `auditoria-detalle`).
        Traerlos aqui haria que listar 16 000 impresiones moviera 750 MB.
        Con `defer` la fila queda en ~200 bytes, asi que el listado completo
        cabe en memoria del navegador y el filtrado es instantaneo, igual que
        el roster de "Imprimir credenciales".

        Los datos personales se resuelven contra el roster (cacheado 5 min):
        el historial guarda num_empleado, no una copia del nombre.
        """
        empleados = self._indice_empleados()

        registros = (
            EnrolamientoCredencial.objects
            .defer('canvas_frente', 'canvas_reverso')
            .order_by('-fecha_registro')
        )

        desde = (request.query_params.get('desde') or '').strip()
        hasta = (request.query_params.get('hasta') or '').strip()
        if desde:
            registros = registros.filter(fecha_registro__date__gte=desde)
        if hasta:
            registros = registros.filter(fecha_registro__date__lte=hasta)

        registros = list(registros)

        # Quien imprimio/modifico cada fila se guarda como IntegerField suelto
        # (id_usuario_registra/id_usuario_modifica), no como FK -- ver
        # AuditoriaUsuarioMixin. Se resuelve a nombre de usuario en UNA sola
        # consulta para todo el listado, no una por fila.
        ids_usuario = {r.id_usuario_registra for r in registros if r.id_usuario_registra}
        ids_usuario |= {r.id_usuario_modifica for r in registros if r.id_usuario_modifica}
        usuarios_por_id = {
            u.id: (u.get_full_name().strip() or u.username)
            for u in Usuario.objects.filter(id__in=ids_usuario)
        } if ids_usuario else {}

        filas = []
        for r in registros:
            empleado = empleados.get((r.num_empleado or '').strip()) or {}
            nombre = ' '.join(filter(None, [
                (empleado.get('nombres') or '').strip(),
                (empleado.get('primer_apellido') or '').strip(),
                (empleado.get('segundo_apellido') or '').strip(),
            ]))
            filas.append({
                'id_enrolamiento': r.id_enrolamiento,
                'num_empleado': r.num_empleado,
                'nombre': nombre,
                'curp': (empleado.get('curp') or '').strip(),
                'area': (empleado.get('area') or '').strip(),
                'folio': r.folio,
                'fecha_expedicion': r.fecha_expedicion,
                'fin_vig': r.fin_vig,
                'plantilla_credencial': r.plantilla_credencial,
                'con_ajustes': bool(r.con_ajustes),
                'fecha_registro': r.fecha_registro,
                'usuario_registra': usuarios_por_id.get(r.id_usuario_registra) or '',
                'fecha_modificacion': r.fecha_modificacion,
                'usuario_modifica': usuarios_por_id.get(r.id_usuario_modifica) or '',
            })

        return Response({
            'status': 'success',
            'total': len(filas),
            'resultados': filas,
        })

    @action(detail=False, methods=['get'], url_path='auditoria-detalle')
    def auditoria_detalle(self, request):
        """
        Lienzos congelados de UNA impresion, para reproducirla en pantalla.

        Se pide por separado del listado justamente porque es lo pesado. Si la
        impresion es anterior a que se guardara el snapshot, `canvas_frente`
        viene vacio y la pantalla lo informa en vez de dibujar una credencial
        reconstruida (que seria una suposicion, no un registro).
        """
        try:
            id_enrolamiento = int(request.query_params.get('id') or 0)
        except (TypeError, ValueError):
            id_enrolamiento = 0

        if not id_enrolamiento:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar id.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registro = EnrolamientoCredencial.objects.filter(
            id_enrolamiento=id_enrolamiento
        ).first()
        if not registro:
            return Response(
                {'status': 'error', 'mensaje': 'No existe esa impresion.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        empleado = self._indice_empleados().get((registro.num_empleado or '').strip()) or {}
        nombre = ' '.join(filter(None, [
            (empleado.get('nombres') or '').strip(),
            (empleado.get('primer_apellido') or '').strip(),
            (empleado.get('segundo_apellido') or '').strip(),
        ]))

        return Response({
            'status': 'success',
            'id_enrolamiento': registro.id_enrolamiento,
            'num_empleado': registro.num_empleado,
            'nombre': nombre,
            'folio': registro.folio,
            'fecha_expedicion': registro.fecha_expedicion,
            'fin_vig': registro.fin_vig,
            'plantilla_credencial': registro.plantilla_credencial,
            'con_ajustes': bool(registro.con_ajustes),
            'fecha_registro': registro.fecha_registro,
            'tiene_snapshot': bool(registro.canvas_frente),
            'canvas_frente': registro.canvas_frente,
            'canvas_reverso': registro.canvas_reverso,
        })

    @action(detail=False, methods=['get'], url_path='auditoria-medios')
    def auditoria_medios(self, request):
        """
        Empleados con MAS DE UNA impresion, para la pestaña "Auditoría" del
        inventario de medios: muestra la foto y la firma vigentes y, detras,
        todas las versiones que se llegaron a imprimir.

        Solo se listan los reimpresos porque son los unicos donde hay algo que
        comparar: con una sola impresion, la version archivada y la vigente
        son la misma.

        La consulta va sobre la columna `medios` (unos cuantos bytes por fila),
        NO sobre los lienzos: recorrer los ~47 KB de cada impresion solo para
        saber que foto se uso haria inservible esta pantalla.
        """
        registros = (
            EnrolamientoCredencial.objects
            .exclude(num_empleado__isnull=True)
            .exclude(num_empleado='')
            .values('num_empleado', 'medios', 'folio', 'fecha_registro', 'id_enrolamiento')
            .order_by('num_empleado', '-fecha_registro')
        )

        por_empleado = {}
        for r in registros:
            por_empleado.setdefault(r['num_empleado'].strip(), []).append(r)

        empleados = self._indice_empleados()
        busqueda = (request.query_params.get('busqueda') or '').strip().upper()

        filas = []
        for numero, impresiones in por_empleado.items():
            if len(impresiones) < 2:
                continue

            datos = empleados.get(numero) or {}
            nombre = ' '.join(filter(None, [
                (datos.get('nombres') or '').strip(),
                (datos.get('primer_apellido') or '').strip(),
                (datos.get('segundo_apellido') or '').strip(),
            ]))

            if busqueda and busqueda not in numero.upper() and busqueda not in nombre.upper():
                continue

            # Versiones DISTINTAS, en orden de la mas reciente a la mas
            # antigua. Se cuentan por ruta: como el archivo se nombra por el
            # hash de su contenido, dos rutas iguales son literalmente la
            # misma imagen, aunque se hayan impreso en fechas distintas.
            versiones = {'foto': [], 'firma': []}
            for imp in impresiones:
                for rol in ('foto', 'firma'):
                    ruta = (imp.get('medios') or {}).get(rol)
                    if ruta and ruta not in versiones[rol]:
                        versiones[rol].append(ruta)

            filas.append({
                'num_empleado': numero,
                'nombre': nombre,
                'curp': (datos.get('curp') or '').strip(),
                'total_impresiones': len(impresiones),
                'ultima_impresion': impresiones[0]['fecha_registro'],
                # Lo vigente en disco hoy, que es lo que se imprimiria ahora.
                'foto_actual': media_utils.url_publica(media_utils.resolver_foto(numero)),
                'firma_actual': media_utils.url_publica(media_utils.resolver_firma(numero)),
                'versiones_foto': len(versiones['foto']),
                'versiones_firma': len(versiones['firma']),
                # Cambio de medios: la version archivada mas reciente no es la
                # unica que existe. Es la señal que hace interesante la fila.
                'cambio_medios': len(versiones['foto']) > 1 or len(versiones['firma']) > 1,
            })

        filas.sort(key=lambda f: (not f['cambio_medios'], -f['total_impresiones']))

        return Response({
            'status': 'success',
            'total': len(filas),
            'resultados': filas,
        })

    @action(detail=False, methods=['get'], url_path='auditoria-medios-detalle')
    def auditoria_medios_detalle(self, request):
        """
        Foto y firma que llevo CADA impresion de un empleado, de la mas
        reciente a la mas antigua, junto con lo que hay hoy en disco.

        Las rutas apuntan al archivo historico, asi que siguen mostrando la
        imagen con la que se expidio esa credencial aunque despues se haya
        recapturado.
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = (
            EnrolamientoCredencial.objects
            .filter(num_empleado=num_empleado)
            .values('id_enrolamiento', 'folio', 'fecha_registro', 'fecha_expedicion', 'medios')
            .order_by('-fecha_registro')
        )

        foto_actual = media_utils.resolver_foto(num_empleado)
        firma_actual = media_utils.resolver_firma(num_empleado)

        # Para saber si una impresion llevaba la imagen que sigue vigente hay
        # que comparar CONTENIDO, no rutas: lo archivado vive en
        # `historico/<hash>.ext` y lo vigente en `fotos/<numero>.ext`, asi que
        # por ruta jamas coincidirian. El hash del archivo vigente se calcula
        # una sola vez y se contrasta con el que ya lleva el nombre archivado.
        vigente = {
            'foto': media_utils.hash_contenido(foto_actual),
            'firma': media_utils.hash_contenido(firma_actual),
        }

        impresiones = []
        for r in registros:
            medios = r.get('medios') or {}
            fila = {
                'id_enrolamiento': r['id_enrolamiento'],
                'folio': r['folio'],
                'fecha_registro': r['fecha_registro'],
                'fecha_expedicion': r['fecha_expedicion'],
            }
            for rol in ('foto', 'firma'):
                ruta = medios.get(rol)
                fila[rol] = media_utils.url_publica(ruta, versionada=False)
                fila[f'{rol}_vigente'] = bool(
                    ruta and vigente[rol] and media_utils.hash_desde_ruta(ruta) == vigente[rol]
                )
            impresiones.append(fila)

        return Response({
            'status': 'success',
            'num_empleado': num_empleado,
            'foto_actual': media_utils.url_publica(foto_actual),
            'firma_actual': media_utils.url_publica(firma_actual),
            'impresiones': impresiones,
        })

    @action(detail=False, methods=['get'], url_path='auditoria-empleado')
    def auditoria_empleado(self, request):
        """
        Todas las impresiones de un empleado, de la mas reciente a la mas
        antigua -- para poder recorrer sus credenciales desde el visor sin
        volver al listado. Sin lienzos, por lo mismo que `auditoria`.
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = (
            EnrolamientoCredencial.objects
            .filter(num_empleado=num_empleado)
            .defer('canvas_frente', 'canvas_reverso')
            .order_by('-fecha_registro')
        )

        return Response({
            'status': 'success',
            'resultados': [{
                'id_enrolamiento': r.id_enrolamiento,
                'folio': r.folio,
                'fecha_expedicion': r.fecha_expedicion,
                'fin_vig': r.fin_vig,
                'plantilla_credencial': r.plantilla_credencial,
                'con_ajustes': bool(r.con_ajustes),
                'fecha_registro': r.fecha_registro,
            } for r in registros],
        })

    @action(detail=False, methods=['get'], url_path='ultima-impresion')
    def ultima_impresion(self, request):
        """
        Ultima credencial impresa de un empleado, para reproducirla al volver
        a seleccionarlo: misma plantilla y, si los hubo, los ajustes de
        posicion que se le hicieron.

        NO devuelve el folio ni la fecha para reutilizarlos -- cada
        reimpresion consume folio nuevo y lleva la fecha del dia. El folio
        anterior queda en el historial, que es justo lo que da la trazabilidad
        de que esa persona tuvo antes otra credencial.
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registro = (
            EnrolamientoCredencial.objects
            .filter(num_empleado=num_empleado)
            # Los lienzos quedan diferidos: Django solo los trae de la BD si
            # abajo se acceden, cosa que pasa unicamente cuando hubo ajustes.
            .defer('canvas_frente', 'canvas_reverso')
            .order_by('-fecha_registro')
            .first()
        )

        if not registro:
            return Response({'status': 'success', 'encontrado': False})

        return Response({
            'status': 'success',
            'encontrado': True,
            'id_enrolamiento': registro.id_enrolamiento,
            'plantilla_credencial': registro.plantilla_credencial,
            'folio_anterior': registro.folio,
            'fecha_expedicion': registro.fecha_expedicion,
            'fecha_registro': registro.fecha_registro,
            # Desde que el snapshot se guarda SIEMPRE, "hay lienzo" ya no
            # significa "lo ajustaron a mano": eso lo dice `con_ajustes`. Si se
            # restaurara el lienzo en toda reimpresion, una credencial impresa
            # limpia dejaria de seguir a su plantilla y los cambios de diseño
            # nunca alcanzarian a las reimpresiones.
            'tiene_ajustes': bool(registro.con_ajustes),
            # El lienzo solo viaja cuando hay ajustes que restaurar. Mandarlo
            # siempre sumaria ~47 KB a CADA seleccion de empleado para que el
            # navegador lo tirara sin usarlo. La auditoria, que si los necesita
            # todos, los pide por `auditoria-detalle`.
            'canvas_frente': registro.canvas_frente if registro.con_ajustes else None,
            'canvas_reverso': registro.canvas_reverso if registro.con_ajustes else None,
            'total_impresiones': EnrolamientoCredencial.objects.filter(
                num_empleado=num_empleado
            ).count(),
        })

    @action(detail=False, methods=['get'], url_path='medios')
    def medios(self, request):
        """
        Resuelve las rutas de foto/firma en disco para un num_empleado y, si
        no hay nada guardado con ese numero, reintenta con el CURP.

        El respaldo por CURP cubre el "enrolamiento previo": es comun capturar
        foto/firma de personal cuyo movimiento de ingreso todavia no se aplica
        en el sistema, por lo que aun no tiene num_empleado asignado. Esas
        capturas se guardan nombradas por CURP y, en cuanto el roster SIG ya
        trae a esa persona con numero, este respaldo permite cruzarlas.

        `foto_origen`/`firma_origen` indican con cual identificador se
        encontro cada archivo ('principal' = num_empleado, 'respaldo' = curp),
        para que el front sepa si al imprimir hay que migrarlos al nombre
        definitivo (ver accion migrar-medios-curp).
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        # El roster SIG solo trae CURP; se acepta rfc por si quien llama ya lo
        # tiene a la mano (ambos comparten el prefijo de 10 caracteres que se
        # usa para el cruce).
        respaldo = (request.query_params.get('curp') or request.query_params.get('rfc') or '').strip()

        if not num_empleado and not respaldo:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado, curp o rfc.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ruta_foto, origen_foto = media_utils.resolver_con_respaldo(
            num_empleado, respaldo, media_utils.carpeta_fotos(), media_utils.EXTENSIONES_FOTO
        )
        ruta_firma, origen_firma = media_utils.resolver_con_respaldo(
            num_empleado, respaldo, media_utils.carpeta_firmas(), media_utils.EXTENSIONES_FIRMA
        )

        origenes_previos = {'respaldo', 'prefijo'}

        return Response({
            'status': 'success',
            'num_empleado': num_empleado,
            'identificador_respaldo': respaldo,
            'foto': media_utils.url_publica(ruta_foto),
            'firma': media_utils.url_publica(ruta_firma),
            'foto_ruta': ruta_foto,
            'firma_ruta': ruta_firma,
            'foto_origen': origen_foto,
            'firma_origen': origen_firma,
            # True si algun medio se encontro por RFC/CURP (enrolamiento
            # previo) y todavia no se renombra al num_empleado definitivo.
            'requiere_migracion': bool(
                num_empleado
                and (origen_foto in origenes_previos or origen_firma in origenes_previos)
            ),
            'encontrado': bool(ruta_foto or ruta_firma),
        })

    @action(detail=True, methods=['post'], url_path='guardar-medios')
    def guardar_medios(self, request, pk=None):
        """
        Guarda foto y/o firma (base64) en MEDIA_ROOT, nombradas por
        num_empleado. No hay ruta que persistir en el modelo -- foto_url/
        firma_url la resuelven solas en cuanto el archivo existe en disco.
        """
        registro = self.get_object()
        identificador = registro.num_empleado or request.data.get('num_empleado')

        if not identificador:
            return Response(
                {'status': 'error', 'mensaje': 'El registro no tiene num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        actualizados = []
        try:
            if request.data.get('foto'):
                media_utils.guardar_foto(request.data['foto'], identificador)
                actualizados.append('foto')

            if request.data.get('firma'):
                media_utils.guardar_firma(request.data['firma'], identificador)
                actualizados.append('firma')
        except media_utils.MediaError as exc:
            return Response(
                {'status': 'error', 'mensaje': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not actualizados:
            return Response(
                {'status': 'error', 'mensaje': 'No se recibio foto ni firma.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registro.fecha_modificacion = timezone.now()
        registro.save(update_fields=['fecha_modificacion'])

        return Response({
            'status': 'success',
            'actualizados': actualizados,
            'foto': registro.foto_url,
            'firma': registro.firma_url,
        })

    @action(detail=False, methods=['post'], url_path='guardar-medios-empleado')
    def guardar_medios_empleado(self, request):
        """
        Guarda foto y/o firma (base64) en MEDIA_ROOT nombradas por
        num_empleado o, si todavia no tiene numero asignado, por RFC.

        No requiere que exista un registro EnrolamientoCredencial -- la
        resolucion de medios es puramente por convencion de nombre de
        archivo (ver media_utils), asi que sirve tanto para un empleado del
        roster SIG (desde "Imprimir credenciales") como para un enrolamiento
        previo por RFC, de personal cuyo movimiento de ingreso aun no se
        aplica y por tanto todavia no tiene num_empleado.
        """
        num_empleado = (request.data.get('num_empleado') or '').strip()
        rfc = (request.data.get('rfc') or request.data.get('curp') or '').strip()
        identificador = num_empleado or rfc

        if not identificador:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado o rfc.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        actualizados = []
        try:
            if request.data.get('foto'):
                media_utils.guardar_foto(request.data['foto'], identificador)
                actualizados.append('foto')

            if request.data.get('firma'):
                media_utils.guardar_firma(request.data['firma'], identificador)
                actualizados.append('firma')
        except media_utils.MediaError as exc:
            return Response(
                {'status': 'error', 'mensaje': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not actualizados:
            return Response(
                {'status': 'error', 'mensaje': 'No se recibio foto ni firma.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'status': 'success',
            'identificador': identificador,
            'num_empleado': num_empleado,
            'rfc': rfc,
            'actualizados': actualizados,
            'foto': media_utils.url_publica(media_utils.resolver_foto(identificador)),
            'firma': media_utils.url_publica(media_utils.resolver_firma(identificador)),
        })

    # Cache del indice del roster. Traer las ~16 400 filas desde la BD remota
    # tarda ~4 s, y el inventario lo necesita completo. Filtrar en SQL con
    # LEFT(CURP,10) IN (...) solo baja a ~2.9 s y ademas NO coincide con
    # prefijo_identidad() -- que normaliza el CURP antes de cortar -- por lo
    # que el inventario marcaria como cruzables registros que luego la
    # migracion no encuentra. Se prefiere la version correcta con un cache
    # corto: el roster solo se actualiza cada 30 min (sync de Celery), asi que
    # 5 minutos de vigencia no puede devolver datos significativamente viejos.
    _CACHE_INDICE_ROSTER: dict = {}
    _CACHE_INDICE_TTL_SEGUNDOS = 300
    # Mismo motivo y misma vigencia, pero indexado por num_empleado: lo usa el
    # inventario de fotos/firmas ya nombradas por numero.
    _CACHE_EMPLEADOS: dict = {}

    def _indice_empleados(self):
        """Mapa num_empleado -> datos del roster, cacheado 5 minutos."""
        ahora = time.monotonic()
        cacheado = EnrolamientoCredencialViewSet._CACHE_EMPLEADOS
        if (
            cacheado.get('indice') is not None
            and (ahora - cacheado.get('marca', 0)) < self._CACHE_INDICE_TTL_SEGUNDOS
        ):
            return cacheado['indice']

        indice = {}
        for fila in SicreTblSig.objects.values(
            'no_empleado', 'nombres', 'primer_apellido', 'segundo_apellido', 'curp', 'area'
        ):
            numero = (fila['no_empleado'] or '').strip()
            if numero:
                indice.setdefault(numero, fila)

        EnrolamientoCredencialViewSet._CACHE_EMPLEADOS = {'indice': indice, 'marca': ahora}
        return indice

    def _indice_roster_por_prefijo(self, refrescar=False):
        """Mapa prefijo de identidad (10 caracteres) -> empleado del roster SIG."""
        ahora = time.monotonic()
        cacheado = EnrolamientoCredencialViewSet._CACHE_INDICE_ROSTER
        if (
            not refrescar
            and cacheado.get('indice') is not None
            and (ahora - cacheado.get('marca', 0)) < self._CACHE_INDICE_TTL_SEGUNDOS
        ):
            return cacheado['indice']

        indice = {}
        filas = (
            SicreTblSig.objects
            .exclude(curp__isnull=True).exclude(curp='')
            .exclude(no_empleado__isnull=True).exclude(no_empleado='')
            .values('curp', 'no_empleado', 'nombres', 'primer_apellido', 'segundo_apellido')
        )
        for fila in filas:
            prefijo = media_utils.prefijo_identidad(fila['curp'])
            if prefijo:
                indice.setdefault(prefijo, fila)

        EnrolamientoCredencialViewSet._CACHE_INDICE_ROSTER = {
            'indice': indice, 'marca': ahora,
        }
        return indice

    @action(detail=False, methods=['get'], url_path='enrolamientos-previos')
    def enrolamientos_previos(self, request):
        """
        Inventario de fotos/firmas nombradas por RFC que todavia no se han
        renombrado a num_empleado.

        No hay tabla en BD que las registre: los archivos en MEDIA_ROOT son la
        unica fuente de verdad, por lo que la lista se construye leyendo el
        directorio.

        A cada archivo se le adjunta `cruce`: el empleado del roster SIG cuyo
        CURP comparte el prefijo de 10 caracteres con ese RFC. Si viene con
        cruce, ya se puede renombrar al num_empleado definitivo; si viene en
        null, esa persona todavia no aparece en el roster (o el nombre del
        archivo es basura).
        """
        registros = media_utils.listar_enrolamientos_previos()
        indice = self._indice_roster_por_prefijo()

        for registro in registros:
            empleado = indice.get(media_utils.prefijo_identidad(registro['rfc']))
            if not empleado:
                registro['cruce'] = None
                continue

            nombre = ' '.join(filter(None, [
                empleado.get('nombres'),
                empleado.get('primer_apellido'),
                empleado.get('segundo_apellido'),
            ])).strip()
            registro['cruce'] = {
                'num_empleado': empleado['no_empleado'],
                'curp': empleado['curp'],
                'nombre': nombre,
            }

        return Response({
            'status': 'success',
            'registros': registros,
            'total': len(registros),
            'cruzables': sum(1 for r in registros if r['cruce']),
        })

    @action(detail=False, methods=['post'], url_path='migrar-medios-lote')
    def migrar_medios_lote(self, request):
        """
        Renombra en bloque las capturas que ya cruzan con el roster.

        Recibe opcionalmente `rfcs` (lista) para limitar el alcance; sin ella
        migra todo lo cruzable. El num_empleado destino NUNCA se toma del
        cliente: se recalcula aqui contra el roster, para que un payload
        manipulado no pueda asignarle la foto de una persona a otra.
        """
        solicitados = request.data.get('rfcs')
        filtro = {str(r).strip().upper() for r in solicitados} if solicitados else None

        indice = self._indice_roster_por_prefijo()
        resultados = []

        for registro in media_utils.listar_enrolamientos_previos():
            rfc = registro['rfc']
            if filtro is not None and rfc.upper() not in filtro:
                continue

            empleado = indice.get(media_utils.prefijo_identidad(rfc))
            if not empleado:
                continue

            migrados = media_utils.migrar_medios(empleado['no_empleado'], rfc)
            if migrados['foto'] or migrados['firma']:
                resultados.append({
                    'rfc': rfc,
                    'num_empleado': empleado['no_empleado'],
                    'migrados': migrados,
                })

        return Response({
            'status': 'success',
            'total_migrados': len(resultados),
            'resultados': resultados,
        })

    # ------------------------------------------------------------------
    # Consecutivo de folio
    # ------------------------------------------------------------------

    CLAVE_FOLIO_POR_DEFECTO = 'credencial'

    def _obtener_consecutivo(self, clave=None):
        registro, _ = ConsecutivoFolio.objects.get_or_create(
            clave=clave or self.CLAVE_FOLIO_POR_DEFECTO
        )
        return registro

    @action(detail=False, methods=['get'], url_path='folio-actual')
    def folio_actual(self, request):
        """Proximo folio a emitir, sin consumirlo."""
        consecutivo = self._obtener_consecutivo(request.query_params.get('clave'))
        return Response({
            'status': 'success',
            'clave': consecutivo.clave,
            'valor': consecutivo.valor,
            'longitud': consecutivo.longitud,
            'folio': consecutivo.formateado(),
        })

    @action(detail=False, methods=['post'], url_path='folio-establecer')
    def folio_establecer(self, request):
        """
        Fija manualmente el proximo folio. Acepta el numero suelto o ya
        formateado con ceros ('000123'); de la longitud del texto recibido se
        toma el relleno, para que el operador pueda cambiar de 6 a 7 digitos
        sin tocar configuracion.
        """
        crudo = str(request.data.get('folio', '')).strip()
        digitos = ''.join(ch for ch in crudo if ch.isdigit())
        if not digitos:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar un folio numerico.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        consecutivo = self._obtener_consecutivo(request.data.get('clave'))
        consecutivo.valor = int(digitos)
        consecutivo.longitud = max(len(digitos), 1)
        consecutivo.id_usuario_modifica = request.data.get('id_usuario')
        consecutivo.save(update_fields=[
            'valor', 'longitud', 'id_usuario_modifica', 'fecha_modificacion',
        ])

        return Response({
            'status': 'success',
            'valor': consecutivo.valor,
            'longitud': consecutivo.longitud,
            'folio': consecutivo.formateado(),
        })

    @action(detail=False, methods=['post'], url_path='folio-consumir')
    def folio_consumir(self, request):
        """
        Entrega el folio actual y avanza el contador, de forma ATOMICA.

        El bloqueo de fila (select_for_update dentro de la transaccion) es lo
        que impide que dos estaciones imprimiendo a la vez reciban el mismo
        folio: la segunda espera a que la primera confirme su incremento.
        Leer y luego escribir sin bloqueo si permitiria esa colision.
        """
        clave = request.data.get('clave') or self.CLAVE_FOLIO_POR_DEFECTO
        # get_or_create fuera de la transaccion: select_for_update exige que la
        # fila ya exista.
        self._obtener_consecutivo(clave)

        with transaction.atomic():
            consecutivo = ConsecutivoFolio.objects.select_for_update().get(clave=clave)
            emitido = consecutivo.formateado()
            consecutivo.valor += 1
            consecutivo.id_usuario_modifica = request.data.get('id_usuario')
            consecutivo.save(update_fields=[
                'valor', 'id_usuario_modifica', 'fecha_modificacion',
            ])
            siguiente = consecutivo.formateado()

        return Response({
            'status': 'success',
            'folio_emitido': emitido,
            'folio_siguiente': siguiente,
            'valor': consecutivo.valor,
            'longitud': consecutivo.longitud,
        })

    @action(detail=False, methods=['post'], url_path='medios-lote')
    def medios_lote(self, request):
        """
        Resuelve foto/firma de varios RFC de golpe.

        Lo usa "Enrolamiento previo" para reconstruir la sesion tras recargar:
        el navegador recuerda que RFC lleva capturados y con esto recupera sus
        URLs. Es deliberadamente barato -- solo toca el sistema de archivos,
        sin cruzar contra el roster -- a diferencia de `enrolamientos-previos`,
        que trae las ~16 400 filas del roster para calcular el cruce.
        """
        rfcs = request.data.get('rfcs') or []
        if not isinstance(rfcs, list):
            return Response(
                {'status': 'error', 'mensaje': 'rfcs debe ser una lista.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = []
        for valor in rfcs:
            rfc = str(valor or '').strip()
            if not rfc:
                continue
            ruta_foto = media_utils.resolver_foto(rfc)
            ruta_firma = media_utils.resolver_firma(rfc)
            # Sin ningun archivo, esa captura ya no existe en disco (la
            # borraron desde el inventario, por ejemplo): se omite para que
            # la sesion no muestre tarjetas fantasma.
            if not ruta_foto and not ruta_firma:
                continue
            registros.append({
                'rfc': rfc,
                'foto': media_utils.url_publica(ruta_foto),
                'firma': media_utils.url_publica(ruta_firma),
            })

        return Response({'status': 'success', 'registros': registros})

    @action(detail=False, methods=['get'], url_path='medios-empleado')
    def medios_empleado(self, request):
        """
        Fotos y firmas ya nombradas por num_empleado, cruzadas con el roster
        para poder buscarlas por nombre.

        Va paginado y con busqueda EN SERVIDOR, a diferencia del inventario de
        pendientes: son ~13 300 archivos, y mandarlos todos obligaria al
        navegador a montar 13 300 <img> de golpe.

        La busqueda cubre num_empleado, nombre, CURP y tambien RFC: aunque el
        roster no guarda RFC, sus 10 primeros caracteres coinciden con los del
        CURP, asi que un RFC tecleado se resuelve por ese prefijo.
        """
        termino = (request.query_params.get('busqueda') or '').strip()
        # Fecha de captura (YYYY-MM-DD). Se compara contra el mtime del
        # archivo, que es cuando se escribio en disco.
        fecha = (request.query_params.get('fecha') or '').strip()
        try:
            pagina = max(1, int(request.query_params.get('pagina', 1)))
            tam = min(200, max(1, int(request.query_params.get('tam_pagina', 60))))
        except ValueError:
            return Response(
                {'status': 'error', 'mensaje': 'pagina y tam_pagina deben ser numeros.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        archivos = media_utils.listar_medios_por_identificador(solo_numericos=True)

        empleados = self._indice_empleados()

        registros = []
        for numero, medios in archivos.items():
            empleado = empleados.get(numero)
            nombre = ''
            if empleado:
                nombre = ' '.join(filter(None, [
                    empleado.get('nombres'),
                    empleado.get('primer_apellido'),
                    empleado.get('segundo_apellido'),
                ])).strip()

            registros.append({
                'num_empleado': numero,
                'nombre': nombre,
                'curp': (empleado or {}).get('curp') or '',
                'area': (empleado or {}).get('area') or '',
                # False = el archivo existe pero esa persona ya no esta en el
                # roster (baja antigua o nombre invalido). Se muestra igual:
                # el punto de esta pantalla es tener control del disco.
                'en_roster': empleado is not None,
                'foto': medios['foto'],
                'firma': medios['firma'],
                'fecha': medios['fecha'],
            })

        if fecha:
            try:
                dia = datetime.strptime(fecha, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'status': 'error', 'mensaje': 'fecha debe venir como YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # El mtime es epoch en hora del servidor; se compara la fecha local
            # para que "6 de agosto" signifique el dia natural del operador y
            # no un tramo corrido en UTC.
            registros = [
                r for r in registros
                if r['fecha'] and datetime.fromtimestamp(r['fecha']).date() == dia
            ]

        if termino:
            busca = termino.upper()
            prefijo = media_utils.prefijo_identidad(termino)
            def coincide(r):
                if busca in r['num_empleado'].upper(): return True
                if busca in r['nombre'].upper(): return True
                if busca in r['curp'].upper(): return True
                # RFC tecleado -> se compara contra el prefijo del CURP.
                return bool(prefijo) and r['curp'].upper().startswith(prefijo)
            registros = [r for r in registros if coincide(r)]

        registros.sort(key=lambda r: (r['nombre'] or '\uffff', r['num_empleado']))

        total = len(registros)
        inicio = (pagina - 1) * tam
        return Response({
            'status': 'success',
            'total': total,
            'pagina': pagina,
            'tam_pagina': tam,
            'total_paginas': max(1, (total + tam - 1) // tam),
            'registros': registros[inicio:inicio + tam],
        })

    @action(detail=False, methods=['post'], url_path='renombrar-medios')
    def renombrar_medios(self, request):
        """
        Cambia el nombre con el que estan guardadas una foto y una firma.

        Caso tipico: el enrolador tecleo mal el RFC y por eso la credencial no
        cruza con sus archivos. Aqui se corrige el nombre y el cruce funciona
        de inmediato, sin volver a tomar la foto.
        """
        actual = (request.data.get('identificador_actual') or '').strip()
        nuevo = (request.data.get('identificador_nuevo') or '').strip()

        try:
            renombrados = media_utils.renombrar_medios(actual, nuevo)
        except media_utils.MediaError as exc:
            return Response(
                {'status': 'error', 'mensaje': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'status': 'success',
            'identificador_actual': actual,
            'identificador_nuevo': nuevo,
            'renombrados': renombrados,
            'foto': media_utils.url_publica(media_utils.resolver_foto(nuevo)),
            'firma': media_utils.url_publica(media_utils.resolver_firma(nuevo)),
        })

    @action(detail=False, methods=['post'], url_path='borrar-medios-previo')
    def borrar_medios_previo(self, request):
        """
        Elimina foto y firma guardadas por RFC. Necesario para corregir una
        captura hecha con un RFC mal escrito, que de otro modo quedaria
        huerfana en disco sin forma de cruzarse nunca.
        """
        rfc = (request.data.get('rfc') or request.data.get('curp') or '').strip()
        if not rfc:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar rfc.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'status': 'success',
            'rfc': rfc,
            'borrados': media_utils.borrar_medios(rfc),
        })

    @action(detail=False, methods=['post'], url_path='migrar-medios')
    def migrar_medios(self, request):
        """
        Renombra foto y firma guardadas por RFC (enrolamiento previo) para que
        pasen a llamarse por num_empleado, ahora que el movimiento de ingreso
        ya se aplico y la persona tiene numero asignado.

        El cruce se hace por el prefijo de 10 caracteres comun a RFC y CURP
        (ver media_utils.resolver_por_prefijo), porque el roster SIG solo
        entrega CURP y la homoclave del RFC no es derivable de el.

        Se invoca al imprimir la credencial, que es el punto en el que el
        cruce ya quedo confirmado. Si ya existe un archivo con el
        num_empleado, ese se respeta y no se sobreescribe.
        """
        num_empleado = (request.data.get('num_empleado') or '').strip()
        respaldo = (request.data.get('curp') or request.data.get('rfc') or '').strip()

        if not num_empleado or not respaldo:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado y curp (o rfc).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        migrados = media_utils.migrar_medios(num_empleado, respaldo)

        return Response({
            'status': 'success',
            'num_empleado': num_empleado,
            'identificador_respaldo': respaldo,
            'migrados': migrados,
            'foto': media_utils.url_publica(media_utils.resolver_foto(num_empleado)),
            'firma': media_utils.url_publica(media_utils.resolver_firma(num_empleado)),
        })

    # El action 'marcar-impreso' se elimino junto con la columna `impreso`:
    # ahora cada fila de la tabla ES una impresion, asi que no hay nada que
    # marcar. Ver registrar-impresion.


class AcuseCredencialViewSet(viewsets.ViewSet):
    """
    Acuses de alta/baja (PDF o imagen del documento firmado), cargados desde
    la pantalla "Acuses" -- hasta AcuseCredencial.MAX_POR_EMPLEADO (10) por
    empleado y tipo, porque un empleado puede tener varios movimientos de
    alta o de baja a lo largo del tiempo.

    Deliberadamente NO es un ModelViewSet: exponer create/update/destroy
    genericos sobre AcuseCredencial permitiria crear una fila de auditoria
    sin el archivo correspondiente (o viceversa). Las operaciones validas son
    `subir`, `eliminar`, `mapa`, `por_empleado`, `auditoria` -- todas pasan
    por media_utils para que archivo y fila de base de datos se mantengan
    juntos.
    """

    @action(detail=False, methods=['get'], url_path='mapa')
    def mapa(self, request):
        """
        {num_empleado: {'alta': N, 'baja': N}} -- cuantos acuses de cada tipo
        tiene YA cada empleado, para pintar la columna "n/10" de la ag-Grid
        de "Acuses" sin disparar una peticion por fila. Se cuenta desde la
        BASE DE DATOS (ya no escaneando el disco como antes de soportar
        varios acuses por empleado): cada fila de AcuseCredencial es un
        archivo real -- ver `subir` -- asi que la tabla es la fuente de
        verdad de cuantos hay, y contar filas es mas barato que listar
        directorios con miles de archivos.
        """
        conteos: dict[str, dict[str, int]] = {}
        agregados = (
            AcuseCredencial.objects.values('num_empleado', 'tipo')
            .annotate(n=Count('id_acuse'))
        )
        for fila in agregados:
            registro = conteos.setdefault(fila['num_empleado'], {'alta': 0, 'baja': 0})
            registro[fila['tipo']] = fila['n']
        return Response({'status': 'success', 'registros': conteos})

    @action(detail=False, methods=['post'], url_path='subir')
    def subir(self, request):
        """
        Guarda el archivo (base64, PDF o imagen) como un acuse NUEVO -- nunca
        reemplaza uno existente, a diferencia del diseno original: ahora un
        empleado puede acumular hasta MAX_POR_EMPLEADO acuses del mismo tipo.
        Abierto a cualquier usuario autenticado con acceso a la pantalla
        (gating por `ver_acuses` en el frontend/ruta) -- es captura operativa
        del dia a dia, no una operacion administrativa, igual que
        `guardar-medios-empleado`.
        """
        num_empleado = (request.data.get('num_empleado') or '').strip()
        tipo = (request.data.get('tipo') or '').strip()
        archivo = request.data.get('archivo')

        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if tipo not in (AcuseCredencial.TIPO_ALTA, AcuseCredencial.TIPO_BAJA):
            return Response(
                {'status': 'error', 'mensaje': "El tipo debe ser 'alta' o 'baja'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not archivo:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar el archivo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # select_for_update() dentro de la transaccion bloquea las filas
        # existentes de este empleado+tipo mientras se decide el proximo
        # numero libre, igual que folio_consumir con ConsecutivoFolio -- sin
        # esto, dos cargas casi simultaneas podrian calcular el mismo hueco y
        # chocar contra la restriccion UNIQUE (num_empleado, tipo, numero).
        with transaction.atomic():
            existentes = set(
                AcuseCredencial.objects.select_for_update()
                .filter(num_empleado=num_empleado, tipo=tipo)
                .values_list('numero', flat=True)
            )
            if len(existentes) >= AcuseCredencial.MAX_POR_EMPLEADO:
                return Response(
                    {
                        'status': 'error',
                        'mensaje': f'Ya se cargaron {AcuseCredencial.MAX_POR_EMPLEADO} acuses de '
                                   f'{tipo} para este empleado. Elimina alguno antes de subir otro.',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            numero = next(n for n in range(1, AcuseCredencial.MAX_POR_EMPLEADO + 1) if n not in existentes)

            try:
                media_utils.guardar_acuse(archivo, num_empleado, tipo, numero)
            except media_utils.MediaError as exc:
                return Response(
                    {'status': 'error', 'mensaje': str(exc)},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            registro = AcuseCredencial.objects.create(
                num_empleado=num_empleado, tipo=tipo, numero=numero,
                id_usuario_carga=_usuario_actual_id(request),
            )

        return Response(
            {
                'status': 'success',
                'id_acuse': registro.id_acuse,
                'num_empleado': num_empleado,
                'tipo': tipo,
                'numero': numero,
                'archivo': registro.archivo_url,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='eliminar')
    def eliminar(self, request):
        """
        Borra UN acuse (archivo + fila) por `id_acuse`. Deja libre su
        `numero` para la proxima carga de ese empleado+tipo -- `subir`
        siempre busca el primer hueco disponible, no el final de la lista.
        """
        id_acuse = request.data.get('id_acuse')
        if not id_acuse:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar id_acuse.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            registro = AcuseCredencial.objects.get(id_acuse=id_acuse)
        except AcuseCredencial.DoesNotExist:
            return Response(
                {'status': 'error', 'mensaje': 'Ese acuse ya no existe.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        media_utils.borrar_acuse(registro.num_empleado, registro.tipo, registro.numero)
        registro.delete()

        return Response({'status': 'success'})

    @action(detail=False, methods=['get'], url_path='por-empleado')
    def por_empleado(self, request):
        """
        Acuses de un empleado -- de alta y baja por omision, o solo un tipo
        si se manda `tipo` (usado por el modal de gestion, que solo necesita
        la lista de UN tipo a la vez).
        """
        num_empleado = (request.query_params.get('num_empleado') or '').strip()
        if not num_empleado:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar num_empleado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = AcuseCredencial.objects.filter(num_empleado=num_empleado)
        tipo = (request.query_params.get('tipo') or '').strip()
        if tipo:
            registros = registros.filter(tipo=tipo)

        return Response({
            'status': 'success',
            'resultados': AcuseCredencialSerializer(registros, many=True).data,
        })

    @action(detail=False, methods=['get'], url_path='auditoria')
    def auditoria(self, request):
        """
        Historial completo de acuses, para la pestaña "Auditoría" de la
        pantalla "Acuses": quien subio cada acuse y, si se reemplazo despues,
        quien y cuando.

        Los nombres (empleado y usuarios) se resuelven en dos consultas
        adicionales, no una por fila -- igual criterio que
        EnrolamientoCredencialViewSet.auditoria(). No hace falta el cache de
        roster de esa clase: aqui como maximo hay 20 filas por empleado
        (hasta 10 de alta y 10 de baja, ver AcuseCredencial.MAX_POR_EMPLEADO),
        muy por debajo de las ~16 mil del roster completo.
        """
        registros = list(AcuseCredencial.objects.all().order_by('-fecha_carga'))

        nums_empleado = {r.num_empleado for r in registros if r.num_empleado}
        empleados = {
            e['no_empleado']: e
            for e in SicreTblSig.objects.filter(no_empleado__in=nums_empleado)
            .values('no_empleado', 'nombres', 'primer_apellido', 'segundo_apellido')
        } if nums_empleado else {}

        ids_usuario = {r.id_usuario_carga for r in registros if r.id_usuario_carga}
        ids_usuario |= {r.id_usuario_modifica for r in registros if r.id_usuario_modifica}
        usuarios_por_id = {
            u.id: (u.get_full_name().strip() or u.username)
            for u in Usuario.objects.filter(id__in=ids_usuario)
        } if ids_usuario else {}

        filas = []
        for r in registros:
            empleado = empleados.get((r.num_empleado or '').strip()) or {}
            nombre = ' '.join(filter(None, [
                (empleado.get('nombres') or '').strip(),
                (empleado.get('primer_apellido') or '').strip(),
                (empleado.get('segundo_apellido') or '').strip(),
            ]))
            filas.append({
                'id_acuse': r.id_acuse,
                'num_empleado': r.num_empleado,
                'nombre': nombre,
                'tipo': r.tipo,
                'numero': r.numero,
                'archivo': r.archivo_url,
                'fecha_carga': r.fecha_carga,
                'usuario_carga': usuarios_por_id.get(r.id_usuario_carga) or '',
                'fecha_modificacion': r.fecha_modificacion,
                'usuario_modifica': usuarios_por_id.get(r.id_usuario_modifica) or '',
            })

        return Response({
            'status': 'success',
            'total': len(filas),
            'resultados': filas,
        })


class PlantillaCredencialViewSet(AuditoriaUsuarioMixin, viewsets.ModelViewSet):
    """CRUD de plantillas disenadas en el editor tipo canvas (Fabric.js)."""
    queryset = PlantillaCredencial.objects.all()
    serializer_class = PlantillaCredencialSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['clave', 'nombre', 'descripcion']

    # Ver y usar plantillas (list/retrieve/por-defecto, que consume "Imprimir
    # credenciales" en cada seleccion) se queda abierto. Solo DISEÑAR una
    # plantilla -- crearla, editarla, borrarla o cambiar cual es la que se usa
    # por omision -- exige `plantillas_administrar`.
    _ACCIONES_ADMINISTRAR = {
        'create', 'update', 'partial_update', 'destroy', 'marcar_por_defecto',
        # Subir/borrar fondos no estaba en este conjunto -- quedaba abierto
        # (AllowAny) igual que list/retrieve, cuando administrar los fondos
        # es la misma operacion de diseño que el resto de este conjunto.
        'subir_fondo', 'borrar_fondo',
    }

    def get_permissions(self):
        if self.action in self._ACCIONES_ADMINISTRAR:
            return [IsAuthenticated(), tiene_permiso('plantillas_administrar')()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        activo = self.request.query_params.get('activo')
        if activo is not None and str(activo) != '':
            qs = qs.filter(activo=str(activo).lower() in ('1', 'true', 'True'))
        return qs

    @action(detail=False, methods=['get'], url_path='por-defecto')
    def por_defecto(self, request):
        """
        Plantilla que se carga por omision en "Imprimir credenciales".

        Si nadie ha marcado una explicitamente, cae a la primera plantilla
        activa para que la pantalla siga siendo usable en vez de aparecer
        vacia.
        """
        plantilla = PlantillaCredencial.objects.filter(
            por_defecto=True, activo=True
        ).first()

        origen = 'marcada'
        if not plantilla:
            plantilla = PlantillaCredencial.objects.filter(activo=True).first()
            origen = 'primera_activa'

        if not plantilla:
            return Response(
                {'status': 'not_found', 'mensaje': 'No hay ninguna plantilla activa.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response({
            'status': 'success',
            'origen': origen,
            'plantilla': PlantillaCredencialSerializer(plantilla).data,
        })

    @action(detail=True, methods=['post'], url_path='marcar-por-defecto')
    def marcar_por_defecto(self, request, pk=None):
        """
        Marca esta plantilla como la de uso por omision y desmarca cualquier
        otra, en una sola transaccion para que nunca queden dos marcadas.
        """
        plantilla = self.get_object()

        with transaction.atomic():
            PlantillaCredencial.objects.filter(por_defecto=True).exclude(
                pk=plantilla.pk
            ).update(por_defecto=False)
            plantilla.por_defecto = True
            plantilla.activo = True  # una plantilla inactiva no puede ser la default
            plantilla.save(update_fields=['por_defecto', 'activo', 'fecha_modificacion'])

        return Response({
            'status': 'success',
            'plantilla': PlantillaCredencialSerializer(plantilla).data,
        })

    @action(detail=False, methods=['post'], url_path='subir-fondo')
    def subir_fondo(self, request):
        """
        Recibe un fondo en base64 y lo guarda en MEDIA_ROOT/plantillas/.

        Se usa desde el editor cuando el usuario sube su propia imagen de fondo
        para el anverso o el reverso.
        """
        contenido = request.data.get('imagen')
        nombre = request.data.get('nombre')

        if not contenido:
            return Response(
                {'status': 'error', 'mensaje': 'No se recibio la imagen.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not nombre:
            nombre = f"fondo_{timezone.now().strftime('%Y%m%d%H%M%S')}"

        try:
            ruta = media_utils.guardar_fondo_plantilla(contenido, nombre)
        except media_utils.MediaError as exc:
            return Response(
                {'status': 'error', 'mensaje': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'status': 'success',
            'ruta': ruta,
            'url': media_utils.url_publica(ruta),
        })

    @action(detail=False, methods=['get'], url_path='fondos-disponibles')
    def fondos_disponibles(self, request):
        """Lista los fondos ya presentes en MEDIA_ROOT/plantillas/."""
        carpeta = Path(settings.MEDIA_ROOT) / settings.MEDIA_DIR_PLANTILLAS
        extensiones = {'.png', '.jpg', '.jpeg', '.webp', '.bmp'}

        archivos = []
        if carpeta.is_dir():
            for archivo in sorted(carpeta.iterdir()):
                if archivo.is_file() and archivo.suffix.lower() in extensiones:
                    ruta = f'{settings.MEDIA_DIR_PLANTILLAS}/{archivo.name}'
                    archivos.append({
                        'nombre': archivo.name,
                        'ruta': ruta,
                        'url': media_utils.url_publica(ruta),
                    })

        return Response({'status': 'success', 'fondos': archivos})

    @action(detail=False, methods=['post'], url_path='borrar-fondo')
    def borrar_fondo(self, request):
        """
        Borra un fondo de MEDIA_ROOT/plantillas/.

        Antes de esto no habia forma de quitar un fondo ya subido -- el
        editor solo permitia agregarlos. Se niega el borrado si ALGUNA
        plantilla (activa o no) todavia lo usa como fondo_frente/fondo_reverso:
        borrarlo de todas formas dejaria esa plantilla apuntando a un archivo
        inexistente, y el editor/preview la mostraria con el fondo en blanco
        sin ninguna pista de por que.
        """
        ruta = (request.data.get('ruta') or '').strip()
        if not ruta:
            return Response(
                {'status': 'error', 'mensaje': 'Debe indicar la ruta del fondo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Solo se puede borrar dentro de la carpeta de fondos -- sin esto,
        # una ruta como '../../settings.py' se resolveria fuera de MEDIA_ROOT.
        prefijo = f'{settings.MEDIA_DIR_PLANTILLAS}/'
        nombre_archivo = ruta[len(prefijo):] if ruta.startswith(prefijo) else ruta
        if not ruta.startswith(prefijo) or '/' in nombre_archivo or '..' in nombre_archivo:
            return Response(
                {'status': 'error', 'mensaje': 'Ruta de fondo invalida.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        en_uso = PlantillaCredencial.objects.filter(
            Q(fondo_frente=ruta) | Q(fondo_reverso=ruta)
        ).values_list('nombre', flat=True)
        if en_uso:
            return Response({
                'status': 'error',
                'mensaje': f'No se puede borrar: lo usan estas plantillas: {", ".join(en_uso)}.',
            }, status=status.HTTP_409_CONFLICT)

        borrado = media_utils.borrar(ruta)
        if not borrado:
            return Response(
                {'status': 'error', 'mensaje': 'No se encontro ese fondo en el servidor.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response({'status': 'success'})

    @action(detail=True, methods=['post'], url_path='duplicar')
    def duplicar(self, request, pk=None):
        """Crea una copia editable de una plantilla existente."""
        original = self.get_object()

        clave_nueva = (request.data.get('clave') or f'{original.clave}_COPIA').strip().upper().replace(' ', '_')
        if PlantillaCredencial.objects.filter(clave=clave_nueva).exists():
            return Response(
                {'status': 'error', 'mensaje': f'Ya existe una plantilla con clave {clave_nueva}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        copia = PlantillaCredencial.objects.create(
            clave=clave_nueva,
            nombre=request.data.get('nombre') or f'{original.nombre} (copia)',
            descripcion=original.descripcion,
            fondo_frente=original.fondo_frente,
            fondo_reverso=original.fondo_reverso,
            canvas_frente=original.canvas_frente,
            canvas_reverso=original.canvas_reverso,
            ancho_px=original.ancho_px,
            alto_px=original.alto_px,
            ancho_mm=original.ancho_mm,
            alto_mm=original.alto_mm,
            activo=True,
        )

        return Response(
            {'status': 'success', 'plantilla': PlantillaCredencialSerializer(copia).data},
            status=status.HTTP_201_CREATED,
        )


class UnidadAdministrativaViewSet(viewsets.ModelViewSet):
    """
    CRUD del catalogo de unidades administrativas.

    `nombre_compactado` es el texto EXACTO que se imprime en el campo de
    adscripcion de la credencial. Hoy las aduanas traen 'ANAM' y las
    Direcciones Generales su acronimo, pero eso es solo el valor con el que se
    sembro el catalogo: si se edita aqui, la credencial cambia de inmediato
    sin tocar codigo -- por ejemplo si mas adelante quieren que una aduana
    imprima su propio nombre.
    """
    queryset = UnidadAdministrativa.objects.all()
    serializer_class = UnidadAdministrativaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'nombre_compactado']
    ordering_fields = ['nombre', 'nombre_compactado', 'activo']
    ordering = ['nombre']
    pagination_class = None   # son ~66 filas: paginar solo estorba

    # Consultar el catalogo (list/retrieve, incluido `areas-sin-catalogo`) se
    # queda abierto -- lo necesita cualquiera que imprima credenciales para
    # ver el nombre compactado. Solo EDITARLO exige `areas_administrar`.
    _ACCIONES_ADMINISTRAR = {'create', 'update', 'partial_update', 'destroy'}

    def get_permissions(self):
        if self.action in self._ACCIONES_ADMINISTRAR:
            return [IsAuthenticated(), tiene_permiso('areas_administrar')()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        activo = self.request.query_params.get('activo')
        if activo in ('0', '1'):
            queryset = queryset.filter(activo=(activo == '1'))
        return queryset

    def list(self, request, *args, **kwargs):
        """
        Lista el catalogo y adjunta a cada unidad cuantos empleados del roster
        le corresponden. Se resuelve con UNA consulta agregada y un cruce en
        memoria: hacerlo por fila serian 66 consultas contra una BD remota.
        """
        unidades = list(self.filter_queryset(self.get_queryset()))

        conteo = {}
        filas = (
            SicreTblSig.objects
            .exclude(area__isnull=True).exclude(area='')
            .values_list('area', flat=True)
        )
        for area in filas:
            clave = UnidadAdministrativa.normalizar(area)
            conteo[clave] = conteo.get(clave, 0) + 1

        for unidad in unidades:
            unidad.total_empleados = conteo.get(unidad.nombre_normalizado, 0)

        serializer = self.get_serializer(unidades, many=True)
        return Response({
            'status': 'success',
            'total': len(unidades),
            'registros': serializer.data,
        })

    @action(detail=False, methods=['get'], url_path='areas-sin-catalogo')
    def areas_sin_catalogo(self, request):
        """
        Areas presentes en el roster que NO tienen entrada en el catalogo.

        Es la alerta que avisa que una unidad nueva llego con el sync y su
        credencial saldria con el nombre largo: sin esto, el hueco solo se
        descubriria al imprimir.
        """
        catalogo = set(
            UnidadAdministrativa.objects.values_list('nombre_normalizado', flat=True)
        )

        faltantes = {}
        filas = (
            SicreTblSig.objects
            .exclude(area__isnull=True).exclude(area='')
            .values_list('area', flat=True)
        )
        for area in filas:
            clave = UnidadAdministrativa.normalizar(area)
            if clave not in catalogo:
                registro = faltantes.setdefault(clave, {'area': area.strip(), 'empleados': 0})
                registro['empleados'] += 1

        return Response({
            'status': 'success',
            'total': len(faltantes),
            'registros': sorted(faltantes.values(), key=lambda r: -r['empleados']),
        })


# ==========================================================================
# Administracion: usuarios, roles y catalogo de permisos
# ==========================================================================
# Gateado a EsSuperusuario en las TRES vistas -- ver docstring de esa clase
# en auth.py: quien puede crear cuentas o repartir permisos debe ser alguien
# de maxima confianza, y `is_superuser` (nativo de Django) ya resuelve ese
# caso sin necesitar una tabla de roles-de-administrador aparte.

Usuario = get_user_model()


class UsuarioViewSet(viewsets.ModelViewSet):
    """
    CRUD de `auth_user` para la pantalla de administracion.

    No hay borrado fisico (`destroy` no esta expuesto vía DRF salvo que se
    quiera; se deja fuera a proposito): un usuario desactivado conserva su
    huella en `id_usuario_registra`/`id_usuario_modifica` de todo lo que
    hizo. Borrarlo de verdad dejaria esas columnas apuntando a un id
    inexistente, y la auditoria dejaria de poder decir quien hizo que.
    """
    queryset = Usuario.objects.all().order_by('username')
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, EsSuperusuario]
    http_method_names = ['get', 'post', 'patch', 'head', 'options']

    def get_queryset(self):
        qs = super().get_queryset().select_related('perfil').prefetch_related('groups')
        busqueda = (self.request.query_params.get('busqueda') or '').strip()
        if busqueda:
            qs = qs.filter(
                Q(username__icontains=busqueda) | Q(email__icontains=busqueda)
                | Q(first_name__icontains=busqueda) | Q(last_name__icontains=busqueda)
            )
        return qs

    @action(detail=True, methods=['post'], url_path='activar')
    def activar(self, request, pk=None):
        usuario = self.get_object()
        usuario.is_active = True
        usuario.save(update_fields=['is_active'])
        return Response({'status': 'success', 'is_active': True})

    @action(detail=True, methods=['post'], url_path='desactivar')
    def desactivar(self, request, pk=None):
        """
        Desactivar, no borrar: `is_active=False` le impide iniciar sesion
        (Django lo revisa solo en `authenticate()`) sin destruir el historial
        de lo que esa cuenta hizo.
        """
        usuario = self.get_object()
        if usuario.pk == request.user.pk:
            return Response(
                {'status': 'error', 'mensaje': 'No puedes desactivar tu propia cuenta.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        usuario.is_active = False
        usuario.save(update_fields=['is_active'])
        return Response({'status': 'success', 'is_active': False})

    @action(detail=True, methods=['post'], url_path='set-password')
    def set_password(self, request, pk=None):
        """
        El superusuario fija una contrasena nueva para la cuenta -- no es un
        flujo de "olvide mi contrasena" con correo (este proyecto no tiene
        servicio de correo configurado), es el equivalente administrativo:
        alguien de confianza la restablece directamente.
        """
        usuario = self.get_object()
        password = (request.data.get('password') or '').strip()
        if len(password) < 8:
            return Response(
                {'status': 'error', 'mensaje': 'La contraseña debe tener al menos 8 caracteres.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        usuario.set_password(password)
        usuario.save(update_fields=['password'])
        # Invalida cualquier sesion anterior: quien tuviera el token viejo
        # (por ejemplo, alguien que ya no deberia tener acceso a esa cuenta)
        # no puede seguir usandolo despues de un cambio de contrasena.
        Token.objects.filter(user=usuario).delete()
        return Response({'status': 'success'})


class RolViewSet(viewsets.ModelViewSet):
    """CRUD de `auth_group`, usados como roles: un rol es un conjunto de permisos."""
    queryset = Group.objects.all().order_by('name')
    serializer_class = RolSerializer
    permission_classes = [IsAuthenticated, EsSuperusuario]

    def get_queryset(self):
        return super().get_queryset().annotate(total_usuarios=Count('user', distinct=True))


class PermisoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Catalogo de permisos disponibles, para construir la matriz de la
    pantalla de roles. Solo lectura: el catalogo se define en
    `RecursoSistema.Meta.permissions` (codigo), no desde la UI.
    """
    serializer_class = PermisoSerializer
    permission_classes = [IsAuthenticated, EsSuperusuario]

    def get_queryset(self):
        return Permission.objects.filter(
            content_type__app_label='credencializacion',
            content_type__model='recursosistema',
        ).order_by('codename')
